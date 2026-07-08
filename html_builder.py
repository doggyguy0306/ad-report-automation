"""
광고 성과 HTML 대시보드 생성 모듈
인터넷 연결 없이 브라우저에서 바로 열 수 있는 자기완결형 HTML
차트: inline SVG (외부 라이브러리 불필요)
"""

import math
import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path

# ── 색상 ────────────────────────────────────
C_NAVER  = '#1B6B3A'
C_GOOGLE = '#4285F4'
C_DA     = '#C0392B'
C_META   = '#1877F2'
C_IG     = '#E1306C'
C_NAVY   = '#1E3A5F'
C_GOLD   = '#F4B400'
C_LIGHT  = '#F4F7FB'
C_BORDER = '#E2E8F0'
C_TEXT   = '#2D3748'
C_MUTED  = '#718096'
C_WHITE  = '#FFFFFF'
C_PCT_BG = '#FFFBEB'

MEDIA_COLORS = {
    'Naver': C_NAVER,
    'Google_SA': C_GOOGLE,
    'Google_DA': C_DA,
}

# ── 포맷 헬퍼 ───────────────────────────────
def _f(n):
    try: return f'{int(float(n)):,}'
    except: return str(n)

def _fp(num, den):
    try:
        d = float(den)
        return f'{float(num)/d*100:.2f}%' if d > 0 else '-'
    except: return '-'

def _pv(num, den):
    try:
        d = float(den)
        return float(num)/d if d > 0 else 0.0
    except: return 0.0

def _cd(conv_df):
    if conv_df is None or conv_df.empty: return {}
    d = {}
    for _, r in conv_df.iterrows():
        d[str(r.get('전환유형',''))] = d.get(str(r.get('전환유형','')), 0) + float(r.get('전환수', 0))
    return d


# ════════════════════════════════════════════
# SVG 차트 생성
# ════════════════════════════════════════════

def _svg_line(dates, values, color, title=''):
    """일별 라인 + 수치 표시 SVG"""
    W, H = 740, 310
    PL, PR, PT, PB = 75, 20, (50 if title else 15), 65
    cw, ch = W - PL - PR, H - PT - PB
    n = len(values)
    if n == 0:
        return f'<svg width="{W}" height="{H}"><text x="50%" y="50%" text-anchor="middle" fill="#aaa">데이터 없음</text></svg>'

    max_v = max(values) * 1.1 if max(values) > 0 else 1
    min_v = max(0, min(values) * 0.9)
    rng = max_v - min_v or 1

    def px(i): return PL + (i / (n - 1)) * cw if n > 1 else PL + cw / 2
    def py(v): return PT + (1 - (v - min_v) / rng) * ch

    # 격자선 + Y레이블 (4단계)
    grid = ''
    for k in range(5):
        gv = min_v + rng * k / 4
        gy = py(gv)
        grid += (f'<line x1="{PL}" y1="{gy:.1f}" x2="{PL+cw}" y2="{gy:.1f}" '
                 f'stroke="#EEEEEE" stroke-width="1"/>'
                 f'<text x="{PL-6}" y="{gy+4:.1f}" text-anchor="end" '
                 f'font-size="10" fill="{C_MUTED}">{_f(gv)}</text>')

    # 면적 채우기
    area_pts = f'{PL:.1f},{PT+ch:.1f}'
    for i, v in enumerate(values): area_pts += f' {px(i):.1f},{py(v):.1f}'
    area_pts += f' {px(n-1):.1f},{PT+ch:.1f}'

    # 선
    line_pts = ' '.join(f'{px(i):.1f},{py(v):.1f}' for i, v in enumerate(values))

    # 포인트 + 수치 레이블
    pts_svg = ''
    for i, v in enumerate(values):
        x, y = px(i), py(v)
        lbl = _f(v)
        pts_svg += (f'<circle cx="{x:.1f}" cy="{y:.1f}" r="4.5" fill="{color}" '
                    f'stroke="white" stroke-width="1.5"/>'
                    f'<text x="{x:.1f}" y="{y-11:.1f}" text-anchor="middle" '
                    f'font-size="10" font-weight="600" fill="{C_TEXT}">{lbl}</text>')

    # X레이블 (최대 8개)
    step = max(1, n // 8)
    x_lbl = ''
    for i in range(0, n, step):
        d_lbl = str(dates[i])[-5:]
        xp = px(i)
        x_lbl += (f'<text x="{xp:.1f}" y="{PT+ch+17}" text-anchor="end" '
                  f'font-size="10" fill="{C_MUTED}" '
                  f'transform="rotate(-35 {xp:.1f} {PT+ch+17})">{d_lbl}</text>')

    axes = (f'<line x1="{PL}" y1="{PT}" x2="{PL}" y2="{PT+ch}" stroke="#CCC" stroke-width="1"/>'
            f'<line x1="{PL}" y1="{PT+ch}" x2="{PL+cw}" y2="{PT+ch}" stroke="#CCC" stroke-width="1"/>')

    title_el = (f'<text x="{W/2}" y="30" text-anchor="middle" font-size="14" '
                f'font-weight="bold" fill="{C_NAVY}">{title}</text>') if title else ''
    return (f'<svg viewBox="0 0 {W} {H}" width="100%" xmlns="http://www.w3.org/2000/svg">'
            f'<rect width="{W}" height="{H}" fill="white" rx="6"/>'
            f'{title_el}'
            f'{grid}{axes}'
            f'<polygon points="{area_pts}" fill="{color}" opacity="0.10"/>'
            f'<polyline points="{line_pts}" fill="none" stroke="{color}" stroke-width="2.5" stroke-linejoin="round"/>'
            f'{pts_svg}{x_lbl}</svg>')


def _svg_dual_line(dates, vals1, color1, label1, vals2, color2, label2, title='',
                   prev_vals1=None, prev_vals2=None):
    """도달수 + 조회수 2개 라인 SVG 차트 (전체 날짜 + 호버 툴팁)
    prev_vals1/prev_vals2: 전월 값 (있으면 파선으로 함께 표시)
    """
    W, H = 780, 360
    PL, PR, PT, PB = 52, 12, (55 if title else 20), 88
    cw, ch = W - PL - PR, H - PT - PB
    n = len(vals1)
    if n == 0:
        return (f'<svg width="{W}" height="{H}"><text x="50%" y="50%" '
                f'text-anchor="middle" fill="#aaa">데이터 없음</text></svg>')

    has_prev_chart = prev_vals1 is not None and len(prev_vals1) > 0
    np_ = len(prev_vals1) if has_prev_chart else 0

    all_vals = list(vals1) + list(vals2)
    if has_prev_chart:
        all_vals += list(prev_vals1)
        if prev_vals2:
            all_vals += list(prev_vals2)
    max_v = max(all_vals) * 1.15 if max(all_vals) > 0 else 1
    rng = max_v or 1

    def px(i): return PL + (i / (n - 1)) * cw if n > 1 else PL + cw / 2
    def px_p(i): return PL + (i / (np_ - 1)) * cw if np_ > 1 else PL + cw / 2
    def py(v): return PT + (1 - v / rng) * ch

    # 격자선 + Y레이블
    grid = ''
    for k in range(5):
        gv = rng * k / 4
        gy = py(gv)
        grid += (f'<line x1="{PL}" y1="{gy:.1f}" x2="{PL+cw}" y2="{gy:.1f}" '
                 f'stroke="#EEEEEE" stroke-width="1"/>'
                 f'<text x="{PL-6}" y="{gy+4:.1f}" text-anchor="end" '
                 f'font-size="10" fill="{C_MUTED}">{_f(gv)}</text>')

    def _line_area(vals, color, opacity, pxfn=None):
        if pxfn is None: pxfn = px
        nn = len(vals)
        area = f'{pxfn(0):.1f},{PT+ch:.1f}'
        for i, v in enumerate(vals): area += f' {pxfn(i):.1f},{py(v):.1f}'
        area += f' {pxfn(nn-1):.1f},{PT+ch:.1f}'
        pts  = ' '.join(f'{pxfn(i):.1f},{py(v):.1f}' for i, v in enumerate(vals))
        return (f'<polygon points="{area}" fill="{color}" opacity="{opacity}"/>'
                f'<polyline points="{pts}" fill="none" stroke="{color}" '
                f'stroke-width="2.5" stroke-linejoin="round"/>')

    def _dashed_line(vals, color, pxfn=None):
        if pxfn is None: pxfn = px_p
        pts = ' '.join(f'{pxfn(i):.1f},{py(v):.1f}' for i, v in enumerate(vals))
        return (f'<polyline points="{pts}" fill="none" stroke="{color}" '
                f'stroke-width="1.8" stroke-dasharray="5,3" stroke-linejoin="round" opacity="0.55"/>')

    # 점 + 호버 영역 (각 포인트마다 투명 rect)
    dots_hover = ''
    for i, (v1, v2) in enumerate(zip(vals1, vals2)):
        x  = px(i)
        y1 = py(v1)
        y2 = py(v2)
        d  = str(dates[i])[-5:]
        # 호버 영역: 세로 띠 (인접 포인트 중간까지)
        left  = px(i-1) if i > 0   else PL
        right = px(i+1) if i < n-1 else PL+cw
        hx    = (left + x) / 2
        hw    = (x + right) / 2 - hx
        dots_hover += (
            f'<circle cx="{x:.1f}" cy="{y1:.1f}" r="4.5" fill="{color1}" stroke="white" stroke-width="1.5"/>'
            f'<circle cx="{x:.1f}" cy="{y2:.1f}" r="4.5" fill="{color2}" stroke="white" stroke-width="1.5"/>'
            f'<rect x="{hx:.1f}" y="{PT}" width="{hw:.1f}" height="{ch}" fill="transparent" '
            f'onmousemove="showIgTip(event,\'{d}\',\'{_f(v1)}\',\'{_f(v2)}\')" '
            f'onmouseleave="hideIgTip()" style="cursor:crosshair"/>'
        )

    # X레이블 — 전체 날짜 표시 (-65도 회전, 9px)
    x_lbl = ''
    for i in range(n):
        xp = px(i)
        x_lbl += (f'<text x="{xp:.1f}" y="{PT+ch+10}" text-anchor="end" '
                  f'font-size="9" fill="{C_MUTED}" '
                  f'transform="rotate(-65 {xp:.1f} {PT+ch+10})">{str(dates[i])[-5:]}</text>')

    axes = (f'<line x1="{PL}" y1="{PT}" x2="{PL}" y2="{PT+ch}" stroke="#CCC" stroke-width="1"/>'
            f'<line x1="{PL}" y1="{PT+ch}" x2="{PL+cw}" y2="{PT+ch}" stroke="#CCC" stroke-width="1"/>')

    # 범례
    lx = PL + 10
    ly = 22
    if has_prev_chart:
        legend = (
            f'<circle cx="{lx+7}" cy="{ly}" r="5" fill="{color1}"/>'
            f'<text x="{lx+16}" y="{ly+4}" font-size="10" font-weight="600" fill="{color1}">이번달 {label1}</text>'
            f'<circle cx="{lx+115}" cy="{ly}" r="5" fill="{color2}"/>'
            f'<text x="{lx+124}" y="{ly+4}" font-size="10" font-weight="600" fill="{color2}">이번달 {label2}</text>'
            f'<line x1="{lx+220}" y1="{ly}" x2="{lx+240}" y2="{ly}" stroke="{color1}" stroke-width="1.8" stroke-dasharray="4,2" opacity="0.6"/>'
            f'<text x="{lx+244}" y="{ly+4}" font-size="10" fill="{C_MUTED}">전월 {label1}</text>'
            f'<line x1="{lx+320}" y1="{ly}" x2="{lx+340}" y2="{ly}" stroke="{color2}" stroke-width="1.8" stroke-dasharray="4,2" opacity="0.6"/>'
            f'<text x="{lx+344}" y="{ly+4}" font-size="10" fill="{C_MUTED}">전월 {label2}</text>'
        )
    else:
        legend = (
            f'<circle cx="{lx+7}" cy="{ly}" r="5" fill="{color1}"/>'
            f'<text x="{lx+16}" y="{ly+4}" font-size="11" font-weight="600" fill="{color1}">{label1}</text>'
            f'<circle cx="{lx+80}" cy="{ly}" r="5" fill="{color2}"/>'
            f'<text x="{lx+89}" y="{ly+4}" font-size="11" font-weight="600" fill="{color2}">{label2}</text>'
        )
    title_svg = ((f'<text x="{W/2}" y="32" text-anchor="middle" font-size="14" '
                  f'font-weight="bold" fill="{C_NAVY}">{title}</text>') if title else '')

    # 전월 파선 (채우기 없이 선만)
    prev_lines = ''
    if has_prev_chart:
        prev_lines = _dashed_line(prev_vals1, color1)
        if prev_vals2:
            prev_lines += _dashed_line(prev_vals2, color2)

    return (f'<svg viewBox="0 0 {W} {H}" width="100%" xmlns="http://www.w3.org/2000/svg">'
            f'<rect width="{W}" height="{H}" fill="white" rx="6"/>'
            f'{title_svg}{legend}{grid}{axes}'
            f'{prev_lines}'
            f'{_line_area(vals1, color1, 0.10)}'
            f'{_line_area(vals2, color2, 0.07)}'
            f'{dots_hover}{x_lbl}</svg>')


def _svg_bar_grouped(group_labels, datasets, title=''):
    """grouped bar chart SVG
    datasets: [(name, [values], color), ...]
    """
    W, H = 740, 310
    PL, PR, PT, PB = 75, 20, (50 if title else 15), 55
    cw, ch = W - PL - PR, H - PT - PB
    n, nd = len(group_labels), len(datasets)
    if n == 0 or nd == 0:
        return ''

    all_vals = [v for _, vals, _ in datasets for v in vals]
    max_v = max(all_vals) * 1.1 if all_vals and max(all_vals) > 0 else 1

    grid = ''
    for k in range(5):
        gv = max_v * k / 4
        gy = PT + ch - (gv / max_v) * ch
        grid += (f'<line x1="{PL}" y1="{gy:.1f}" x2="{PL+cw}" y2="{gy:.1f}" '
                 f'stroke="#EEEEEE" stroke-width="1"/>'
                 f'<text x="{PL-6}" y="{gy+4:.1f}" text-anchor="end" '
                 f'font-size="10" fill="{C_MUTED}">{_f(gv)}</text>')

    group_w = cw / n
    bar_w   = group_w * 0.7 / nd
    pad     = group_w * 0.15
    bars    = ''

    for gi, g_lbl in enumerate(group_labels):
        gx = PL + gi * group_w + pad
        for di, (_, vals, color) in enumerate(datasets):
            v  = vals[gi] if gi < len(vals) else 0
            bh = (v / max_v) * ch if max_v > 0 else 0
            bx = gx + di * bar_w
            by = PT + ch - bh
            bars += (f'<rect x="{bx:.1f}" y="{by:.1f}" width="{bar_w:.1f}" '
                     f'height="{bh:.1f}" fill="{color}" rx="3"/>'
                     f'<text x="{bx + bar_w/2:.1f}" y="{max(PT+5, by-4):.1f}" '
                     f'text-anchor="middle" font-size="9" font-weight="bold" fill="{color}">{_f(v)}</text>')
        # X label
        lx = PL + gi * group_w + group_w / 2
        bars += (f'<text x="{lx:.1f}" y="{PT+ch+18}" text-anchor="middle" '
                 f'font-size="11" fill="{C_TEXT}">{g_lbl}</text>')

    # 범례
    legend = ''
    total_leg_w = nd * 110
    lx0 = (W - total_leg_w) / 2
    for di, (name, _, color) in enumerate(datasets):
        lx = lx0 + di * 110
        legend += (f'<rect x="{lx:.1f}" y="{H-20}" width="12" height="12" fill="{color}" rx="2"/>'
                   f'<text x="{lx+16:.1f}" y="{H-10}" font-size="11" fill="{C_TEXT}">{name}</text>')

    axes = (f'<line x1="{PL}" y1="{PT}" x2="{PL}" y2="{PT+ch}" stroke="#CCC" stroke-width="1"/>'
            f'<line x1="{PL}" y1="{PT+ch}" x2="{PL+cw}" y2="{PT+ch}" stroke="#CCC" stroke-width="1"/>')

    title_el = (f'<text x="{W/2}" y="30" text-anchor="middle" font-size="14" '
                f'font-weight="bold" fill="{C_NAVY}">{title}</text>') if title else ''
    return (f'<svg viewBox="0 0 {W} {H}" width="100%" xmlns="http://www.w3.org/2000/svg">'
            f'<rect width="{W}" height="{H}" fill="white" rx="6"/>'
            f'{title_el}'
            f'{grid}{axes}{bars}{legend}</svg>')


def _svg_pie(labels, values, colors, title=''):
    """파이 차트 SVG"""
    W, H = 420, 310
    total = sum(values) if values else 1
    if total == 0: return ''
    cx, cy, r = W * 0.40, H * 0.55, min(W * 0.38, H * 0.42)

    slices, leg = '', ''
    angle = -math.pi / 2

    for i, (lbl, v, color) in enumerate(zip(labels, values, colors)):
        pct  = v / total
        a2   = angle + 2 * math.pi * pct
        x1 = cx + r * math.cos(angle); y1 = cy + r * math.sin(angle)
        x2 = cx + r * math.cos(a2);   y2 = cy + r * math.sin(a2)
        large = 1 if pct > 0.5 else 0
        slices += (f'<path d="M{cx:.1f},{cy:.1f} L{x1:.1f},{y1:.1f} '
                   f'A{r},{r} 0 {large},1 {x2:.1f},{y2:.1f} Z" '
                   f'fill="{color}" stroke="white" stroke-width="2"/>')
        if pct > 0.04:
            mid_a = (angle + a2) / 2
            lx = cx + r * 0.62 * math.cos(mid_a)
            ly = cy + r * 0.62 * math.sin(mid_a)
            slices += (f'<text x="{lx:.1f}" y="{ly:.1f}" text-anchor="middle" '
                       f'font-size="11" font-weight="bold" fill="white">{pct*100:.1f}%</text>')
        # 범례
        ly_l = 55 + i * 24
        leg += (f'<rect x="{W*0.80:.1f}" y="{ly_l}" width="12" height="12" fill="{color}" rx="2"/>'
                f'<text x="{W*0.80+16:.1f}" y="{ly_l+10}" font-size="11" fill="{C_TEXT}">{lbl[:10]}</text>')
        angle = a2

    title_el = (f'<text x="{cx}" y="28" text-anchor="middle" font-size="14" '
                f'font-weight="bold" fill="{C_NAVY}">{title}</text>') if title else ''
    return (f'<svg viewBox="0 0 {W} {H}" width="100%" xmlns="http://www.w3.org/2000/svg">'
            f'<rect width="{W}" height="{H}" fill="white" rx="6"/>'
            f'{title_el}'
            f'{slices}{leg}</svg>')


# ════════════════════════════════════════════
# HTML 섹션 빌더
# ════════════════════════════════════════════

def _insights_html(
    total_노출, total_클릭, total_비용, total_CTR, total_CPC,
    naver_클릭, naver_비용, naver_CTR, naver_CPC,
    gsa_클릭, gsa_비용, gsa_CTR, gsa_CPC,
    gda_클릭, gda_비용, gda_CTR, gda_CPC,
    has_prev,
    p_노출=0, p_클릭=0, p_비용=0,
    p_total_CTR=0.0, p_total_CPC=0.0
) -> str:
    """데이터 기반 시사점 자동 생성 → HTML 반환"""

    goods = []   # ✅ 잘된 점
    warns = []   # ⚠️ 주의사항
    recs  = []   # 💡 실행 제안

    def pct_chg(curr, prev):
        try:
            p = float(prev)
            return (float(curr) - p) / abs(p) * 100 if p != 0 else None
        except Exception:
            return None

    # ── 1. 전월 대비 MoM 분석 ──────────────────
    if has_prev:
        # CTR MoM
        ctr_d = pct_chg(total_CTR, p_total_CTR)
        if ctr_d is not None:
            if ctr_d >= 10:
                goods.append(
                    f'CTR 전월 대비 <strong>+{ctr_d:.1f}%</strong> 향상 — 소재·타겟팅 품질 개선 긍정 신호'
                )
            elif ctr_d <= -10:
                warns.append(
                    f'CTR 전월 대비 <strong>{ctr_d:.1f}%</strong> 하락 — 경쟁 심화 또는 소재 노후화 점검 필요'
                )

        # CPC MoM
        cpc_d = pct_chg(total_CPC, p_total_CPC)
        if cpc_d is not None:
            if cpc_d >= 20:
                warns.append(
                    f'평균 CPC 전월 대비 <strong>+{cpc_d:.1f}%</strong> 상승 — 입찰 경쟁 심화 또는 품질지수 저하 가능성'
                )
                recs.append(
                    '비용 효율 낮은 키워드 정기 점검·정리 필요, 광고 품질지수(QS) 향상 통해 CPC 낮추는 전략 검토 필요'
                )
            elif cpc_d <= -10:
                goods.append(
                    f'평균 CPC 전월 대비 <strong>{cpc_d:.1f}%</strong> 하락 — 클릭당 비용 효율 개선, 키워드 관리 효과 확인'
                )

        # 노출 vs 클릭 성장세
        imp_d = pct_chg(total_노출, p_노출)
        clk_d = pct_chg(total_클릭, p_클릭)
        if imp_d is not None and clk_d is not None:
            if imp_d > 10 and clk_d < -5:
                warns.append(
                    f'노출 <strong>+{imp_d:.1f}%</strong> 증가 vs 클릭 <strong>{clk_d:.1f}%</strong> 감소 — 소재 클릭 유도력 점검 필요'
                )
            elif imp_d > 5 and clk_d is not None and clk_d > 0:
                goods.append(
                    f'노출(+{imp_d:.1f}%) · 클릭(+{clk_d:.1f}%) 동반 증가 — 전반적 성장세 유지 중'
                )

        # 광고비 대비 클릭 효율
        cost_d = pct_chg(total_비용, p_비용)
        if cost_d is not None and clk_d is not None:
            if cost_d > 10 and clk_d is not None and clk_d < 0:
                warns.append(
                    f'광고비 <strong>+{cost_d:.1f}%</strong> 증가 대비 클릭 감소 — 예산 배분 재검토 필요'
                )

    # ── 2. CTR 절대값 평가 ─────────────────────
    # 한국 디지털 광고 업계: SA 3~6%, DA 0.3~1% / 전체 혼합 기준 2~4% 정상
    if total_CTR >= 4.5:
        goods.append(
            f'전체 CTR <strong>{total_CTR:.2f}%</strong> — 업계 평균(2~4%) 상회, 소재·타겟팅 효과적'
        )
    elif total_CTR < 1.5:
        warns.append(
            f'전체 CTR <strong>{total_CTR:.2f}%</strong> — 낮은 수준, 광고 문구·이미지 소재 전면 점검 권장'
        )
        recs.append(
            'SA 광고문구에 구체적 혜택·CTA 추가 필요, DA 배너 소재 최소 분기 1회 교체 검토 필요'
        )

    # ── 3. 매체별 예산 편중 분석 ──────────────
    if total_비용 > 0:
        n_pct = naver_비용 / total_비용 * 100
        g_pct = gsa_비용  / total_비용 * 100
        d_pct = gda_비용  / total_비용 * 100

        top = max([('네이버', n_pct), ('구글 SA', g_pct), ('구글 DA', d_pct)],
                  key=lambda x: x[1])
        if top[1] > 70:
            warns.append(
                f'광고비 <strong>{top[1]:.0f}%</strong>가 {top[0]}에 집중 — 매체 장애·정책 변경 시 리스크, 다각화 검토 필요'
            )

        # CTR 기준 최고/최저 채널 비교
        ch = [('네이버', naver_CTR, naver_비용), ('구글 SA', gsa_CTR, gsa_비용), ('구글 DA', gda_CTR, gda_비용)]
        ch_with_spend = [(n, c, s) for n, c, s in ch if s > 0]
        if len(ch_with_spend) >= 2:
            best  = max(ch_with_spend, key=lambda x: x[1])
            worst = min(ch_with_spend, key=lambda x: x[1])
            if best[0] != worst[0] and worst[1] > 0:
                ratio = best[1] / worst[1] if worst[1] > 0 else 0
                if ratio >= 2:
                    recs.append(
                        f'CTR 최고 매체 <strong>{best[0]}({best[1]:.2f}%)</strong> vs 최저 <strong>{worst[0]}({worst[1]:.2f}%)</strong> — {worst[0]} 예산 일부 {best[0]}으로 이동 검토 필요'
                    )

    # ── 4. CPC 절대값 평가 (한국 업계 기준) ──
    # 검색광고 평균 CPC 300~1000원 / DA 50~300원 범위
    if total_CPC > 1000:
        warns.append(
            f'평균 CPC <strong>₩{_f(total_CPC)}</strong> — 고비용 키워드 위주 운영 가능성, 키워드 구성 점검 필요'
        )
        recs.append(
            '롱테일 키워드 비중 확대 필요 — CPC 절감 + 전환 의도 높은 트래픽 유입 기대'
        )
    elif total_CPC < 100 and total_클릭 > 100:
        goods.append(
            f'평균 CPC <strong>₩{_f(total_CPC)}</strong> — 매우 낮은 수준, 비용 효율적 운영 유지 중'
        )

    # ── 5. 구글 SA vs DA 비교 시사점 ──────────
    if gsa_비용 > 0 and gda_비용 > 0 and gsa_CPC > 0 and gda_CPC > 0:
        if gsa_CPC > gda_CPC * 3:
            recs.append(
                f'구글 SA CPC(₩{_f(gsa_CPC)}) DA(₩{_f(gda_CPC)}) 대비 {gsa_CPC/gda_CPC:.1f}배 — 브랜드 인지 목적 시 DA 비중 확대 검토 필요'
            )

    # ── 6. 항상 제안: 더 나아가기 ────────────
    recs.append(
        '리타겟팅 캠페인 추가 검토 필요 — 브랜드 인지 사용자 재접촉으로 전환율·ROI 향상 기대'
    )

    # ── HTML 조립 ─────────────────────────────
    if not goods and not warns and not recs:
        return ''

    def _block(title, items, cls):
        if not items:
            return ''
        lis = ''.join(f'<li>{i}</li>' for i in items)
        return (
            f'<div class="ins-block {cls}">'
            f'<div class="ins-title">{title}</div>'
            f'<ul>{lis}</ul>'
            f'</div>'
        )

    blocks = (
        _block('✅ 잘된 점',   goods, 'ins-good') +
        _block('⚠️ 주의사항', warns, 'ins-warn') +
        _block('💡 실행 제안', recs,  'ins-rec')
    )
    return f'<div class="ins-grid">{blocks}</div>'


def _pct_mom(curr, prev):
    """전월 대비 % 변화 → (표시문자열, 색상코드). 데이터 없으면 None"""
    try:
        c, p = float(curr), float(prev)
        if p == 0:
            return None
        pct = (c - p) / abs(p) * 100
        sign = '+' if pct >= 0 else ''
        color = '#16A34A' if pct >= 0 else '#DC2626'
        return (f'{sign}{pct:.1f}%', color)
    except Exception:
        return None


def _mini_bar_svg(prev_val, curr_val, color, fmt='n'):
    """전월 vs 당월 비교 미니 막대그래프 SVG (카드 내 삽입용)"""
    W, H   = 156, 68
    BAR_W  = 46
    GAP    = 18
    PT     = 10   # top padding (value label space)
    PB     = 20   # bottom padding (axis label space)
    PL     = (W - BAR_W * 2 - GAP) // 2
    CH     = H - PT - PB   # chart area height
    BOTTOM = PT + CH

    pv = float(prev_val) if prev_val else 0
    cv = float(curr_val) if curr_val else 0
    mx = max(pv, cv, 1)

    ph = max(4, int(pv / mx * CH))
    ch = max(4, int(cv / mx * CH))

    px = PL
    cx = PL + BAR_W + GAP

    def _fmt(v):
        if fmt == 'pct':   return f'{v:.1f}%'
        if fmt == 'won':
            if v >= 1_000_000: return f'₩{v/10000:.0f}만'
            if v >= 10_000:    return f'₩{v/10000:.1f}만'
            return f'₩{int(v):,}'
        if v >= 1_000_000: return f'{v/10000:.0f}만'
        if v >= 10_000:    return f'{v/10000:.1f}만'
        if v >= 1_000:     return f'{v/1000:.1f}k'
        return str(int(v))

    GRAY = '#CBD5E0'
    svg = (
        f'<svg width="{W}" height="{H}" xmlns="http://www.w3.org/2000/svg" '
        f'style="display:block;overflow:visible;flex-shrink:0">'
        # 구분선
        f'<line x1="0" y1="{BOTTOM}" x2="{W}" y2="{BOTTOM}" stroke="#E2E8F0" stroke-width="1"/>'
        # 전월 bar (gray)
        f'<rect x="{px}" y="{BOTTOM - ph}" width="{BAR_W}" height="{ph}" fill="{GRAY}" rx="3"/>'
        # 당월 bar (accent)
        f'<rect x="{cx}" y="{BOTTOM - ch}" width="{BAR_W}" height="{ch}" fill="{color}" rx="3" opacity="0.82"/>'
        # 값 레이블
        f'<text x="{px + BAR_W//2}" y="{BOTTOM - ph - 3}" '
        f'text-anchor="middle" font-size="8.5" fill="#A0AEC0">{_fmt(pv)}</text>'
        f'<text x="{cx + BAR_W//2}" y="{BOTTOM - ch - 3}" '
        f'text-anchor="middle" font-size="8.5" font-weight="700" fill="{color}">{_fmt(cv)}</text>'
        # 축 레이블
        f'<text x="{px + BAR_W//2}" y="{H - 2}" '
        f'text-anchor="middle" font-size="9" fill="#A0AEC0">전월</text>'
        f'<text x="{cx + BAR_W//2}" y="{H - 2}" '
        f'text-anchor="middle" font-size="9" font-weight="600" fill="{color}">당월</text>'
        f'</svg>'
    )
    return svg


def _kpi_card(label, value, sub='', color=C_NAVY, mom=None):
    mom_html = ''
    if mom:
        mom_str, mom_color = mom
        mom_html = (f'<div class="kpi-mom" style="color:{mom_color}">'
                    f'전월 동기간 대비 {mom_str}</div>')
    return (f'<div class="kpi-card" style="border-top:4px solid {color}">'
            f'<div class="kpi-label">{label}</div>'
            f'<div class="kpi-value" style="color:{color}">{value}</div>'
            f'<div class="kpi-sub">{sub}</div>'
            f'{mom_html}'
            f'</div>')


def _kpi_card_expandable(label, value, sub, color, mom, detail_rows, card_id,
                         bar_data=None, prev_label='전월'):
    """상세보기 토글이 있는 KPI 카드.
    detail_rows: [(매체명, 값, sub문자열, color, mom_tuple_or_None), ...]
    bar_data: (prev_val, curr_val, fmt) — 미니 막대그래프 표시용 (전월 데이터 있을 때만)
    """
    mom_html = ''
    if mom:
        mom_str, mom_color = mom
        mom_html = (f'<div class="kpi-mom" style="color:{mom_color}">'
                    f'{prev_label} 동기간 대비 {mom_str}</div>')

    # 미니 막대그래프 (전월 데이터 있을 때만)
    chart_html = ''
    if bar_data is not None:
        prev_val, curr_val, fmt = bar_data
        if float(prev_val) > 0 or float(curr_val) > 0:
            chart_html = _mini_bar_svg(prev_val, curr_val, color, fmt)

    # 매체별 상세 행
    detail_html = ''
    for (m_label, m_val, m_sub, m_color, m_mom) in detail_rows:
        m_mom_html = ''
        if m_mom:
            mm_str, mm_color = m_mom
            m_mom_html = (f'<span class="kpi-dr-mom" style="color:{mm_color}">'
                          f'{prev_label} {mm_str}</span>')
        detail_html += (
            f'<div class="kpi-detail-row">'
            f'<span class="kpi-dr-label" style="color:{m_color}">{m_label}</span>'
            f'<span class="kpi-dr-val">{m_val}</span>'
            f'<span class="kpi-dr-sub">{m_sub}</span>'
            f'{m_mom_html}'
            f'</div>'
        )

    # 텍스트(좌) + 차트(우) 가로 배치
    if chart_html:
        inner = (
            f'<div style="display:flex;align-items:center;gap:10px;min-height:72px">'
            f'  <div style="flex:1;min-width:0">'
            f'    <div class="kpi-label">{label}</div>'
            f'    <div class="kpi-value" style="color:{color}">{value}</div>'
            f'    <div class="kpi-sub">{sub}</div>'
            f'    {mom_html}'
            f'  </div>'
            f'  <div style="flex-shrink:0">{chart_html}</div>'
            f'</div>'
        )
    else:
        inner = (
            f'<div class="kpi-label">{label}</div>'
            f'<div class="kpi-value" style="color:{color}">{value}</div>'
            f'<div class="kpi-sub">{sub}</div>'
            f'{mom_html}'
        )

    return (
        f'<div class="kpi-expand-wrap">'
        f'  <div class="kpi-card" style="border-top:4px solid {color}">'
        f'    {inner}'
        f'    <button class="kpi-detail-btn" onclick="toggleKpi(\'{card_id}\')">'
        f'      <span id="{card_id}-icon">▼</span> 매체별 상세</button>'
        f'  </div>'
        f'  <div id="{card_id}" class="kpi-detail-panel">'
        f'    {detail_html}'
        f'  </div>'
        f'</div>'
    )


def _table_html(headers, rows, stripe=True):
    """범용 HTML 테이블"""
    ths = ''.join(f'<th>{h}</th>' for h in headers)
    trs = ''
    for i, row in enumerate(rows):
        cls = ' class="stripe"' if stripe and i % 2 == 1 else ''
        tds = ''.join(
            f'<td class="pct">{c}</td>' if isinstance(c, str) and c.endswith('%')
            else f'<td>{c}</td>'
            for c in row
        )
        trs += f'<tr{cls}>{tds}</tr>'
    return f'<table><thead><tr>{ths}</tr></thead><tbody>{trs}</tbody></table>'


def _section(title, color, content):
    return (f'<div class="section">'
            f'<div class="section-header" style="background:{color}">{title}</div>'
            f'<div class="section-body">{content}</div>'
            f'</div>')


def _chart_block(title, svg):
    return (f'<div class="chart-block">'
            f'<div class="chart-title">{title}</div>'
            f'{svg}'
            f'</div>')


# ════════════════════════════════════════════
# SNS 채널 데이터 읽기 / HTML 섹션
# ════════════════════════════════════════════

# 채널별 메타 정보
_SNS_META = {
    '📱 인스타그램': {
        'color': '#E1306C', 'emoji': '📱',
        'table_cols': ['날짜', '팔로워수', '게시물수(당일)', '좋아요수', '댓글수', '도달수'],
        'col_idx':    [0,      1,           3,               5,         6,        7],
        'main_col': 1,  # 팔로워수
        'ad_table_cols': ['날짜', '광고비(원)', '광고노출수', '광고클릭수', '광고도달수', '전환수', 'CTR(%)', 'CPC(원)', 'CPM(원)'],
        'ad_col_idx':    [0,      11,           12,           13,           14,           15,       16,       17,       18],
    },
    '🧵 쓰레드': {
        'color': '#444444', 'emoji': '🧵',
        'table_cols': ['날짜', '팔로워수', '게시물수\n(당일)', '조회수', '좋아요수', '댓글수'],
        'col_idx':    [0,      1,           3,                  5,        6,          7],
        'main_col': 1,
    },
    '📝 티스토리': {
        'color': '#FF6600', 'emoji': '📝',
        'table_cols': ['날짜', '방문자수(UV)', '페이지뷰(PV)', '게시물수\n(당일)', '댓글수', '공감수'],
        'col_idx':    [0,      1,              2,               3,                  5,        6],
        'main_col': 1,  # 방문자수
    },
    '💬 카카오톡채널': {
        'color': '#C8A800', 'emoji': '💬',
        'table_cols': ['날짜', '친구수(누적)', '신규친구수', '메시지발송수', '메시지열람수', '열람율(%)'],
        'col_idx':    [0,      1,              2,             4,              5,              6],
        'main_col': 1,  # 친구수
    },
    '▶ 유튜브': {
        'color': '#CC0000', 'emoji': '▶',
        'table_cols': ['날짜', '구독자수', '조회수', '시청시간(시간)', '좋아요수', '댓글수'],
        'col_idx':    [0,      1,           3,        4,                5,          6],
        'main_col': 1,  # 구독자수
    },
}

_SNS_CHANNEL_ORDER = [
    '📱 인스타그램', '🧵 쓰레드', '📝 티스토리', '💬 카카오톡채널', '▶ 유튜브'
]


def _read_sns_channel_data(sns_path: str) -> dict:
    """SNS 관리대장 xlsx 에서 각 채널의 입력 데이터를 읽어 반환.
    반환:
        { channel_name: {'headers': [...], 'rows': [[date, v1, v2, ...], ...]} }
    """
    result = {}
    try:
        wb = openpyxl.load_workbook(sns_path, data_only=True)
        for ch in _SNS_CHANNEL_ORDER:
            if ch not in wb.sheetnames:
                continue
            ws = wb[ch]
            # 헤더 row4 (인스타그램은 광고 컬럼까지)
            max_col = 19 if ch == '📱 인스타그램' else 11
            raw_hdrs = [ws.cell(4, c).value for c in range(1, max_col + 1)]
            headers  = [str(h).replace('\n', '') if h else '' for h in raw_hdrs]
            # 데이터 row5~94: 날짜(col A)가 있는 행만
            # 인스타그램은 광고 컬럼(L~S)까지 읽음
            max_col = 19 if ch == '📱 인스타그램' else 11
            rows = []
            for r in range(5, 95):
                date_val = ws.cell(r, 1).value
                if date_val is None:
                    continue
                row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
                rows.append(row_vals)
            result[ch] = {'headers': headers, 'rows': rows}
    except Exception:
        pass
    return result


def _sns_mom_for_channel(rows, main_col):
    """SNS rows 내에서 전월 동기간 메인 지표 값 반환 (없으면 None)"""
    try:
        from dateutil.relativedelta import relativedelta
        filled = [(pd.Timestamp(r[0]), float(r[main_col]))
                  for r in rows if r[main_col] is not None and r[0] is not None]
        if not filled:
            return None
        curr_end_date = filled[-1][0]
        prev_end_date = curr_end_date - relativedelta(months=1)
        # 전월 동일 날짜에서 가장 가까운 값
        prev_val = None
        min_diff = None
        for d, v in filled:
            diff = abs((d - prev_end_date).days)
            if min_diff is None or diff < min_diff:
                min_diff = diff
                prev_val = v
        return prev_val if min_diff is not None and min_diff <= 5 else None
    except Exception:
        return None


def _sns_summary_cards(channel_data: dict) -> str:
    """채널별 메인 지표 KPI 카드 5개 (한 줄) + 전월 동기간 대비"""
    cards = ''
    for ch in _SNS_CHANNEL_ORDER:
        meta     = _SNS_META[ch]
        color    = meta['color']
        rows     = channel_data.get(ch, {}).get('rows', [])
        main_col = meta['main_col']

        if rows:
            filled = [r for r in rows if main_col < len(r) and r[main_col] is not None]
            latest = filled[-1][main_col] if filled else None
            prev   = filled[-2][main_col] if len(filled) >= 2 else None
            val_str = _f(latest) if latest is not None else '-'
            if prev is not None and latest is not None:
                try:
                    diff = float(latest) - float(prev)
                    sign = '+' if diff >= 0 else ''
                    delta_str   = f'{sign}{_f(diff)}'
                    delta_color = '#16A34A' if diff >= 0 else '#DC2626'
                except Exception:
                    delta_str, delta_color = '-', C_MUTED
            else:
                delta_str, delta_color = '입력 대기', C_MUTED

            # 전월 동기간 대비
            prev_mom_val = _sns_mom_for_channel(rows, main_col)
            mom          = _pct_mom(latest, prev_mom_val) if (latest is not None and prev_mom_val) else None
            mom_html = (f'<div class="kpi-mom" style="color:{mom[1]}">전월 동기간 {mom[0]}</div>'
                        if mom else '')
        else:
            val_str, delta_str, delta_color, mom_html = '미입력', '-', C_MUTED, ''

        headers     = channel_data.get(ch, {}).get('headers', [])
        metric_name = headers[main_col] if headers and len(headers) > main_col else '주요지표'

        cards += (
            f'<div class="sns-kpi" style="border-top:4px solid {color}">'
            f'<div class="sns-ch-name" style="color:{color}">{ch}</div>'
            f'<div class="sns-metric-label">{metric_name}</div>'
            f'<div class="sns-value" style="color:{color}">{val_str}</div>'
            f'<div class="sns-delta" style="color:{delta_color}">전일대비 {delta_str}</div>'
            f'{mom_html}'
            f'</div>'
        )
    return f'<div class="sns-kpi-grid">{cards}</div>'


def _sns_stats_summary(ch: str, data: dict) -> str:
    """채널 기간 통계 (누적 합계 / 일평균) 요약 스트립"""
    meta     = _SNS_META[ch]
    rows     = data.get('rows', [])
    col_idx  = meta['col_idx']
    tbl_cols = meta['table_cols']
    main_col = meta['main_col']
    color    = meta['color']

    if not rows:
        return ''

    filled_rows = [r for r in rows if any(r[ci] is not None for ci in col_idx[1:] if ci < len(r))]
    if not filled_rows:
        return ''

    # 기간
    try:
        first_date = pd.Timestamp(filled_rows[0][0]).strftime('%m/%d')
        last_date  = pd.Timestamp(filled_rows[-1][0]).strftime('%m/%d')
        period_str = f'{first_date} ~ {last_date} ({len(filled_rows)}일)'
    except Exception:
        period_str = f'{len(filled_rows)}일'

    # 메인 지표 증감 (첫 입력값 → 최신값)
    main_vals = [r for r in rows if main_col < len(r) and r[main_col] is not None]
    if len(main_vals) >= 2:
        try:
            diff = float(main_vals[-1][main_col]) - float(main_vals[0][main_col])
            sign = '+' if diff >= 0 else ''
            growth_str  = f'{sign}{_f(diff)}'
            growth_color = '#16A34A' if diff >= 0 else '#DC2626'
        except Exception:
            growth_str, growth_color = '-', C_MUTED
    else:
        growth_str, growth_color = '-', C_MUTED

    main_label = tbl_cols[col_idx.index(main_col)].replace('\n', '')

    boxes = [
        f'<div class="stat-box"><div class="stat-label">집계 기간</div>'
        f'<div class="stat-val">{period_str}</div></div>',
        f'<div class="stat-box"><div class="stat-label">{main_label} 증감</div>'
        f'<div class="stat-val" style="color:{growth_color}">{growth_str}</div>'
        f'<div class="stat-sub">(기간 첫날 대비)</div></div>',
    ]

    # 전월 동기간 데이터 추출
    try:
        from dateutil.relativedelta import relativedelta
        curr_dates = [pd.Timestamp(r[0]) for r in filled_rows if r[0] is not None]
        if curr_dates:
            c_start = min(curr_dates)
            c_end   = max(curr_dates)
            p_start = c_start - relativedelta(months=1)
            p_end   = c_end   - relativedelta(months=1)
            prev_filled = [r for r in rows
                           if r[0] is not None and p_start <= pd.Timestamp(r[0]) <= p_end
                           and any(r[ci] is not None for ci in col_idx[1:] if ci < len(r))]
        else:
            prev_filled = []
    except Exception:
        prev_filled = []

    # 나머지 지표 합계 + 일평균 + 전월 대비 (메인 컬럼 제외)
    for i, ci in enumerate(col_idx):
        if ci == 0 or ci == main_col:
            continue
        vals = [float(r[ci]) for r in rows if ci < len(r) and r[ci] is not None]
        if not vals:
            continue
        col_name = tbl_cols[i].replace('\n', ' ')
        total = sum(vals)
        avg   = total / len(vals)

        # 전월 동기간 합계
        prev_vals = [float(r[ci]) for r in prev_filled if ci < len(r) and r[ci] is not None]
        mom = _pct_mom(total, sum(prev_vals)) if prev_vals else None
        mom_html = (f'<div class="stat-sub" style="color:{mom[1]}">전월 {mom[0]}</div>'
                    if mom else '')

        boxes.append(
            f'<div class="stat-box">'
            f'<div class="stat-label">{col_name}</div>'
            f'<div class="stat-val">{_f(total)}</div>'
            f'<div class="stat-sub">일평균 {_f(avg)}</div>'
            f'{mom_html}'
            f'</div>'
        )

    return f'<div class="stats-strip">{"".join(boxes)}</div>'


def _sns_mini_svg_from_pairs(valid_pairs, n, color):
    """(index, value) 쌍 목록으로 미니 라인 차트 SVG 생성"""
    if len(valid_pairs) < 2:
        return ''
    W, H = 700, 120
    PL, PR, PT, PB = 10, 10, 20, 20
    cw, ch_h = W-PL-PR, H-PT-PB
    vmin = min(v for _, v in valid_pairs)
    vmax = max(v for _, v in valid_pairs)
    rng  = vmax - vmin or 1
    def px(i): return PL + (i/(n-1))*cw if n > 1 else PL+cw/2
    def py(v): return PT + (1-(v-vmin)/rng)*ch_h
    pts      = ' '.join(f'{px(i):.1f},{py(v):.1f}' for i,v in valid_pairs)
    area_pts = f'{px(valid_pairs[0][0]):.1f},{PT+ch_h:.1f} {pts} {px(valid_pairs[-1][0]):.1f},{PT+ch_h:.1f}'
    li, lv   = valid_pairs[-1]
    return (
        f'<svg viewBox="0 0 {W} {H}" width="100%" style="display:block;margin-bottom:8px">'
        f'<rect width="{W}" height="{H}" fill="white"/>'
        f'<polygon points="{area_pts}" fill="{color}" opacity="0.12"/>'
        f'<polyline points="{pts}" fill="none" stroke="{color}" stroke-width="2.5" stroke-linejoin="round"/>'
        f'<circle cx="{px(li):.1f}" cy="{py(lv):.1f}" r="5" fill="{color}" stroke="white" stroke-width="2"/>'
        f'<text x="{px(li):.1f}" y="{max(PT+2, py(lv)-8):.1f}" text-anchor="middle" '
        f'font-size="11" font-weight="bold" fill="{color}">{_f(lv)}</text>'
        f'</svg>'
    )


def _fmt_sns_cell(ci, v):
    """SNS 셀 값 포맷"""
    if v is None:
        return '-'
    if ci == 0:
        try: return pd.Timestamp(v).strftime('%Y-%m-%d')
        except: return str(v)
    elif ci == 16:  # CTR(%)
        try: return f'{float(v):.2f}%'
        except: return str(v)
    elif isinstance(v, float) and v == int(v):
        return f'{int(v):,}'
    elif isinstance(v, (int, float)):
        return _f(v)
    return str(v)


def _sns_channel_detail(ch: str, data: dict, block_id: str = 'sns0') -> str:
    """채널 1개 — 통계 요약 + 트렌드 차트 + 전체 데이터 테이블 (+ 광고 성과)
    block_id: 일별/주별 토글 DOM ID 구분용
    """
    meta     = _SNS_META[ch]
    color    = meta['color']
    rows     = data.get('rows', [])
    col_idx  = meta['col_idx']
    tbl_hdrs = meta['table_cols']
    main_col = meta['main_col']

    if not rows:
        return (
            f'<div style="text-align:center;padding:20px;color:{C_MUTED};font-size:12px">'
            f'아직 데이터가 입력되지 않았습니다.<br>'
            f'<b>SNS_채널_관리대장.xlsx</b> 의 {ch} 탭에 날짜와 수치를 입력하세요.</div>'
        )

    # ── 통계 요약 ────────────────────────────
    stats_html = _sns_stats_summary(ch, data)

    # ── 값이 입력된 행만 (전체 표시) ──────────
    filled_rows = [r for r in rows if any(r[ci] is not None for ci in col_idx[1:] if ci < len(r))]
    display_rows = filled_rows if filled_rows else rows
    tbl_hdrs_clean = [h.replace('\n', ' ') for h in tbl_hdrs]

    # ── 일별 트렌드 차트 ──────────────────────
    trend_vals = []
    for r in display_rows:
        v = r[main_col] if main_col < len(r) else None
        try:    trend_vals.append(float(v)) if v is not None else trend_vals.append(None)
        except: trend_vals.append(None)

    valid_pairs_d = [(i, v) for i, v in enumerate(trend_vals) if v is not None]
    mini_svg_daily = _sns_mini_svg_from_pairs(valid_pairs_d, len(display_rows), color)

    # ── 일별 테이블 (data-date 속성 → 기간 필터 JS 연동) ─────
    daily_tbody = ''
    for i, r in enumerate(reversed(display_rows)):
        date_str = ''
        try:
            dv = r[col_idx[0]] if col_idx[0] < len(r) else None
            date_str = pd.Timestamp(dv).strftime('%Y-%m-%d') if dv else ''
        except Exception:
            pass
        stripe = ' class="stripe daily-row"' if i % 2 == 0 else ' class="daily-row"'
        cells = ''.join(f'<td>{_fmt_sns_cell(ci, r[ci] if ci < len(r) else None)}</td>'
                        for ci in col_idx)
        daily_tbody += f'<tr{stripe} data-date="{date_str}">{cells}</tr>'

    daily_tbl_html = (
        f'<div style="max-height:360px;overflow-y:auto;border-radius:6px">'
        f'<table><thead><tr>'
        + ''.join(f'<th>{h}</th>' for h in tbl_hdrs_clean)
        + f'</tr></thead><tbody>{daily_tbody}</tbody></table></div>'
    )

    # ── 주별 집계 ─────────────────────────────
    weekly_svg = ''
    weekly_tbl_html = '<div style="color:#aaa;font-size:12px;padding:12px">주별 집계 데이터가 없습니다.</div>'
    try:
        df_sns = pd.DataFrame([
            {f'c{ci}': (r[ci] if ci < len(r) else None) for ci in col_idx}
            | {'_date': pd.Timestamp(r[col_idx[0]]) if r[col_idx[0]] else pd.NaT}
            for r in display_rows
        ]).dropna(subset=['_date'])
        df_sns['_week'] = df_sns['_date'].dt.to_period('W')

        agg_dict = {}
        for ci in col_idx[1:]:
            agg_dict[f'c{ci}'] = 'last' if ci == main_col else 'sum'

        weekly = df_sns.groupby('_week', as_index=False).agg(agg_dict)
        weekly['_wstr'] = weekly['_week'].apply(
            lambda p: f"{p.start_time.strftime('%m/%d')}~{p.end_time.strftime('%m/%d')}"
        )

        # 주별 차트
        w_vals = [float(wr[f'c{main_col}']) if pd.notna(wr[f'c{main_col}']) else None
                  for _, wr in weekly.iterrows()]
        w_valid = [(i, v) for i, v in enumerate(w_vals) if v is not None]
        weekly_svg = _sns_mini_svg_from_pairs(w_valid, len(weekly), color)

        # 주별 테이블
        w_hdrs = ['주간'] + tbl_hdrs_clean[1:]
        w_tbody = ''
        for i, (_, wr) in enumerate(weekly.iterrows()):
            stripe = ' class="stripe"' if i % 2 == 0 else ''
            cells = f'<td>{wr["_wstr"]}</td>'
            for ci in col_idx[1:]:
                v = wr.get(f'c{ci}')
                cells += f'<td>{_f(v) if pd.notna(v) else "-"}</td>'
            w_tbody += f'<tr{stripe}>{cells}</tr>'

        weekly_tbl_html = (
            f'<div style="max-height:360px;overflow-y:auto;border-radius:6px">'
            f'<table><thead><tr>'
            + ''.join(f'<th>{h}</th>' for h in w_hdrs)
            + f'</tr></thead><tbody>{w_tbody}</tbody></table></div>'
        )
    except Exception:
        pass

    # ── 집계 토글 + 조합 ─────────────────────
    tbl_section = f'''
<div class="aggr-toggle">
  <button class="aggr-btn aggr-active" id="aggr-daily-{block_id}"
    onclick="setAggr('{block_id}','daily')">일별</button>
  <button class="aggr-btn" id="aggr-weekly-{block_id}"
    onclick="setAggr('{block_id}','weekly')">주별</button>
</div>
<div id="daily-view-{block_id}">
  {mini_svg_daily}
  <div class="sub-title">일별 데이터</div>
  {daily_tbl_html}
</div>
<div id="weekly-view-{block_id}" style="display:none">
  {weekly_svg}
  <div class="sub-title">주별 데이터</div>
  {weekly_tbl_html}
</div>
'''

    # ── 인스타그램 광고 성과 (있는 경우) ──────
    ad_html = ''
    if 'ad_col_idx' in meta:
        ad_col_idx  = meta['ad_col_idx']
        ad_tbl_cols = meta['ad_table_cols']
        ad_filled   = [r for r in rows if len(r) > 11 and
                       any(r[ci] is not None for ci in ad_col_idx[1:] if ci < len(r))]
        if ad_filled:
            ad_tbl_rows = []
            for r in reversed(ad_filled):
                ad_row = [_fmt_sns_cell(ci, r[ci] if ci < len(r) else None)
                          for ci in ad_col_idx]
                ad_tbl_rows.append(ad_row)
            ad_hdrs_clean = [h.replace('\n', ' ') for h in ad_tbl_cols]
            ad_html = (
                f'<div style="margin-top:20px">'
                f'<div style="font-size:13px;font-weight:700;color:{color};'
                f'margin-bottom:8px;padding-bottom:6px;border-bottom:2px solid #FCE4EC">'
                f'📢 광고 성과</div>'
                f'<div style="max-height:300px;overflow-y:auto;border-radius:6px">'
                f'{_table_html(ad_hdrs_clean, ad_tbl_rows)}'
                f'</div></div>'
            )
        else:
            ad_html = (
                f'<div style="margin-top:20px;padding:14px;background:#FFF5F7;'
                f'border-radius:8px;border:1px dashed #F48FB1;font-size:12px;color:{C_MUTED}">'
                f'📢 <b>광고 성과</b> — SNS 관리대장 인스타그램 탭의 L~S열에 광고 데이터를 입력하면 여기에 표시됩니다.'
                f'</div>'
            )

    return stats_html + tbl_section + ad_html


def _ig_insights_html(ig_media_df) -> str:
    """Instagram 게시물 성과 기반 콘텐츠 시사점 자동 생성 → HTML"""
    if ig_media_df is None or ig_media_df.empty or len(ig_media_df) < 2:
        return ''

    from collections import Counter
    df = ig_media_df.copy()

    goods = []
    warns = []
    recs  = []

    # 숫자 정규화
    for col in ['도달수', '조회수', '좋아요', '댓글', '저장수']:
        if col not in df.columns:
            df[col] = 0
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # ── 1. 유형별 평균 성과 비교 ────────────────
    if '유형' in df.columns:
        type_grp = df.groupby('유형').agg(
            평균조회수=('조회수', 'mean'),
            평균도달수=('도달수', 'mean'),
            게시물수=('도달수', 'count')
        ).reset_index()

        v_row = type_grp[type_grp['유형'] == 'VIDEO']
        i_row = type_grp[type_grp['유형'] == 'IMAGE']

        if not v_row.empty and not i_row.empty:
            v_avg = v_row.iloc[0]['평균조회수']
            i_avg = i_row.iloc[0]['평균조회수']
            v_cnt = int(v_row.iloc[0]['게시물수'])
            i_cnt = int(i_row.iloc[0]['게시물수'])
            if v_avg > 0 and i_avg > 0:
                ratio = v_avg / i_avg
                if ratio >= 1.5:
                    goods.append(
                        f'동영상 평균 조회수({_f(int(v_avg))}) 이미지 대비 <strong>{ratio:.1f}배</strong> 높음 — 릴스·숏폼 집중 전략 유효'
                    )
                    recs.append(
                        f'동영상({v_cnt}개) 비중 확대 필요 — 이미지({i_cnt}개) 일부 릴스 포맷 전환 검토'
                    )
                elif ratio < 0.8:
                    goods.append(
                        f'이미지 평균 조회수({_f(int(i_avg))})가 동영상 대비 높음 — 카드뉴스·이미지 포맷 강점 확인'
                    )

    # ── 2. 상위 게시물 캡션 패턴 분석 ────────────
    sort_col = '조회수' if df['조회수'].sum() > 0 else '도달수'
    top_df   = df.nlargest(min(3, len(df)), sort_col)

    def _caption_patterns(caption):
        cap = str(caption)
        pats = []
        if cap.startswith('"') or cap.startswith('“'):
            pats.append('대화체')
        if '...' in cap or '…' in cap:
            pats.append('말줄임(궁금증)')
        if '?' in cap:
            pats.append('질문형')
        if any(k in cap for k in ['사연', '후기', '경험', '이불킥', '실수', '얘기', '썰']):
            pats.append('스토리텔링')
        if any(k in cap for k in ['😢', '🥲', '😅', '😂', '❤️', '🔥', '😱']):
            pats.append('감성이모지')
        if any(k in cap for k in ['혜택', '할인', '이벤트', '공지', '소개', '출시']):
            pats.append('프로모션')
        return pats

    all_pats = []
    for _, row in top_df.iterrows():
        all_pats.extend(_caption_patterns(row.get('캡션', '')))
    dominant = [p for p, c in Counter(all_pats).most_common(2) if c >= 2]

    if dominant:
        goods.append(
            f'상위 게시물 공통 패턴: <strong>{"·".join(dominant)}</strong> — 해당 스타일 지속 활용 권장'
        )

    best = top_df.iloc[0]
    best_val     = int(best.get(sort_col, 0))
    best_caption = str(best.get('캡션', ''))[:22]
    best_type    = str(best.get('유형', '')).replace('VIDEO','동영상').replace('IMAGE','이미지').replace('CAROUSEL_ALBUM','카루셀')
    if best_val > 0:
        goods.append(
            f'최고 성과 {best_type}: "<strong>{best_caption}…</strong>" — {sort_col} <strong>{_f(best_val)}</strong>'
        )

    # ── 3. 저장률 분석 (정보성 지표) ─────────────
    total_saved = df['저장수'].sum()
    total_reach = df['도달수'].sum()
    if total_reach > 0 and total_saved > 0:
        save_rate = total_saved / total_reach * 100
        if save_rate >= 1.0:
            goods.append(
                f'평균 저장률 <strong>{save_rate:.2f}%</strong> — 정보·참고가치 높은 콘텐츠로 인식됨'
            )
        elif save_rate < 0.2:
            warns.append(
                f'저장률 <strong>{save_rate:.2f}%</strong> — 낮은 수준, 캡션 내 "저장해두세요" CTA 추가 검토 필요'
            )

    # ── 4. 댓글 참여 분석 ────────────────────────
    avg_comments = df['댓글'].mean()
    if avg_comments < 0.5 and len(df) >= 3:
        warns.append(
            f'평균 댓글 <strong>{avg_comments:.1f}개</strong>/게시물 — 질문형 마무리 또는 댓글 이벤트로 참여 유도 검토 필요'
        )
    elif avg_comments >= 3:
        goods.append(
            f'평균 댓글 <strong>{avg_comments:.1f}개</strong>/게시물 — 활발한 댓글 참여 유지 중'
        )

    # ── 5. 콘텐츠 방향 제안 ──────────────────────
    if '유형' in df.columns:
        video_ratio = len(df[df['유형'] == 'VIDEO']) / len(df) * 100
        if video_ratio < 30:
            recs.append(
                f'동영상 비중 현재 <strong>{video_ratio:.0f}%</strong> — 릴스(15~30초) 비중 확대 권장, 알고리즘 노출 유리'
            )

    if '대화체' in dominant or '스토리텔링' in dominant:
        recs.append(
            '고객 실제 경험·대화체 캡션 포맷 지속 활용 필요 — 공감 기반 콘텐츠가 도달수·조회수 상위권 차지'
        )
    else:
        recs.append(
            '캡션에 고객 관점 대화체·사연 포맷 도입 검토 필요 — 공감·호기심 유발로 자연 도달 확대 기대'
        )

    if not goods and not warns and not recs:
        return ''

    def _ig_block(title, items, bg, border, title_color):
        if not items:
            return ''
        lis = ''.join(
            f'<li style="margin-bottom:6px;font-size:13px;line-height:1.6">{i}</li>'
            for i in items
        )
        return (
            f'<div style="background:{bg};border:1px solid {border};border-radius:8px;'
            f'padding:14px 16px;flex:1;min-width:220px">'
            f'<div style="font-size:12px;font-weight:700;color:{title_color};margin-bottom:10px">{title}</div>'
            f'<ul style="margin:0;padding-left:18px;color:{C_TEXT}">{lis}</ul>'
            f'</div>'
        )

    blocks = (
        _ig_block('✅ 잘된 점',      goods, '#F0FDF4', '#86EFAC', '#16A34A') +
        _ig_block('⚠️ 주의사항',    warns, '#FFFBEB', '#FCD34D', '#D97706') +
        _ig_block('💡 콘텐츠 방향',  recs,  '#FFF0F5', '#F9A8D4', C_IG)
    )
    return (
        f'<div style="background:#fff;border:1px solid #f0b8d4;border-radius:10px;'
        f'padding:16px;margin-top:16px">'
        f'<div style="font-size:13px;font-weight:600;color:{C_IG};margin-bottom:12px">'
        f'📌 콘텐츠 성과 시사점</div>'
        f'<div style="display:flex;gap:12px;flex-wrap:wrap">{blocks}</div>'
        f'</div>'
    )


def _build_instagram_section(ig_account_df, ig_media_df, prev_ig_media_df=None) -> str:
    """Instagram 유기 성과 섹션 HTML
    ig_account_df 없어도 ig_media_df 만으로 일별 집계를 자동 도출해 차트 표시.
    prev_ig_media_df: 전월 비교 기간 게시물 데이터 (MoM 표시용)
    """
    parts = []

    # ── 게시물 데이터에서 일별 집계 자동 도출 ─────
    if ig_account_df is None and ig_media_df is not None and not ig_media_df.empty:
        _tmp = ig_media_df.copy()
        _tmp['날짜'] = pd.to_datetime(_tmp['날짜']).dt.normalize()
        _agg = {'도달수': 'sum'}
        if '조회수' in _tmp.columns:
            _agg['조회수'] = 'sum'
        daily = _tmp.groupby('날짜', as_index=False).agg(_agg).sort_values('날짜')
        # 시작~종료일 사이 모든 날짜를 채워 연속 표시 (게시물 없는 날 = 0)
        full_range = pd.date_range(start=daily['날짜'].min(), end=daily['날짜'].max(), freq='D')
        daily = daily.set_index('날짜').reindex(full_range, fill_value=0).reset_index()
        daily.rename(columns={'index': '날짜'}, inplace=True)
        ig_account_df = daily

    # ── 전월 지표 사전 계산 ─────────────────────────
    has_prev_ig = prev_ig_media_df is not None and not prev_ig_media_df.empty
    prev_ig_account_df = None
    if has_prev_ig:
        p_post_count  = len(prev_ig_media_df)
        p_total_likes = int(prev_ig_media_df['좋아요'].sum()) if '좋아요' in prev_ig_media_df.columns else 0
        p_total_saved = int(prev_ig_media_df['저장수'].sum()) if '저장수' in prev_ig_media_df.columns else 0
        p_total_views = int(prev_ig_media_df['조회수'].sum()) if '조회수' in prev_ig_media_df.columns else 0
        # 전월 도달수 + 조회수 일별 집계
        _p_tmp = prev_ig_media_df.copy()
        _p_tmp['날짜'] = pd.to_datetime(_p_tmp['날짜']).dt.normalize()
        _p_agg = {'도달수': 'sum'}
        if '조회수' in _p_tmp.columns:
            _p_agg['조회수'] = 'sum'
        p_daily = _p_tmp.groupby('날짜', as_index=False).agg(_p_agg).sort_values('날짜')
        # 전월 날짜 범위 채우기
        if not p_daily.empty:
            _p_range = pd.date_range(start=p_daily['날짜'].min(), end=p_daily['날짜'].max(), freq='D')
            p_daily = p_daily.set_index('날짜').reindex(_p_range, fill_value=0).reset_index()
            p_daily.rename(columns={'index': '날짜'}, inplace=True)
            prev_ig_account_df = p_daily
        p_total_reach = p_daily['도달수'].sum() if '도달수' in p_daily.columns else 0
        _p_posting = p_daily[p_daily['도달수'] > 0] if '도달수' in p_daily.columns else pd.DataFrame()
        p_avg_reach = _p_posting['도달수'].mean() if not _p_posting.empty else 0
    else:
        p_post_count = p_total_likes = p_total_saved = p_total_views = 0
        p_total_reach = p_avg_reach = 0

    # ── 월 레이블 도출 ────────────────────────────
    def _ig_month_label(df):
        if df is None or df.empty: return None
        try:
            return f"{pd.to_datetime(df['날짜'].min()).month}월"
        except: return None
    curr_label = _ig_month_label(ig_media_df) or '이번달'
    prev_label = (_ig_month_label(prev_ig_media_df) if has_prev_ig else None) or '전월'

    def _ig_mom_badge(curr, prev):
        """전월 대비 뱃지 HTML"""
        if not has_prev_ig or prev == 0:
            return ''
        try:
            pct = (float(curr) - float(prev)) / abs(float(prev)) * 100
            sign, clr = ('▲', '#16A34A') if pct >= 0 else ('▼', '#DC2626')
            return (f'<div style="font-size:10px;font-weight:600;color:{clr};margin-top:3px">'
                    f'{sign}{abs(pct):.1f}% <span style="color:#aaa;font-weight:400">전월 대비</span></div>')
        except Exception:
            return ''

    # ── 계정 인사이트 요약 ──────────────────────────
    if ig_account_df is not None and not ig_account_df.empty:
        total_reach = ig_account_df['도달수'].sum() if '도달수' in ig_account_df.columns else 0
        _posting = ig_account_df[ig_account_df['도달수'] > 0] if '도달수' in ig_account_df.columns else pd.DataFrame()
        avg_reach   = _posting['도달수'].mean() if not _posting.empty else 0
        max_day     = ig_account_df.loc[ig_account_df['도달수'].idxmax()] if total_reach > 0 else None

        # 게시물 데이터에서 추가 지표 계산
        post_count   = len(ig_media_df) if ig_media_df is not None and not ig_media_df.empty else 0
        total_likes  = int(ig_media_df['좋아요'].sum()) if ig_media_df is not None and '좋아요' in ig_media_df.columns else 0
        total_saved  = int(ig_media_df['저장수'].sum()) if ig_media_df is not None and '저장수' in ig_media_df.columns else 0
        total_views  = int(ig_media_df['조회수'].sum()) if ig_media_df is not None and '조회수' in ig_media_df.columns else 0
        # 팔로워 수: CSV에 컬럼 있으면 사용
        follower_cnt = int(ig_media_df['팔로워수'].iloc[0]) if ig_media_df is not None and '팔로워수' in ig_media_df.columns else None

        def _ig_info_card(label, value, sub=''):
            return (f'<div style="background:#FDF2F8;border:1px solid #F0B8D4;'
                    f'border-radius:10px;padding:16px;text-align:center;flex:1">'
                    f'<div style="font-size:11px;color:{C_IG};font-weight:600;margin-bottom:6px">{label}</div>'
                    f'<div style="font-size:24px;font-weight:700;color:#333">{value}</div>'
                    f'<div style="font-size:11px;color:#888;margin-top:4px">{sub}</div>'
                    f'</div>')

        # [A] 상단 info 카드 2개
        info_follower = (
            _ig_info_card('팔로워 수', _f(follower_cnt), '현재 기준') if follower_cnt is not None
            else _ig_info_card('총 저장수', _f(total_saved), '기간 합산')
        )
        info_max_day = _ig_info_card(
            '최고 도달일',
            max_day['날짜'].strftime('%m/%d') if max_day is not None else '-',
            f'{_f(max_day["도달수"])}명 도달' if max_day is not None else ''
        )
        parts.append(
            f'<div style="display:flex;gap:14px;margin-bottom:16px">'
            f'{info_follower}{info_max_day}'
            f'</div>'
        )

        # [B] 이번달 vs 전월 비교 바차트 2개 나란히
        if has_prev_ig:
            # 참여 지표
            eng_chart = _svg_bar_grouped(
                ['게시물 수', '좋아요', '저장수'],
                [(curr_label, [float(post_count), float(total_likes), float(total_saved)], C_IG),
                 (prev_label, [float(p_post_count), float(p_total_likes), float(p_total_saved)], '#F0B8D4')],
            )
            # 도달 지표
            if total_views > 0 or p_total_views > 0:
                reach_labels = ['총 도달수', '일평균 도달수', '총 조회수']
                reach_curr_vals = [float(total_reach), float(avg_reach), float(total_views)]
                reach_prev_vals = [float(p_total_reach), float(p_avg_reach), float(p_total_views)]
            else:
                reach_labels = ['총 도달수', '일평균 도달수']
                reach_curr_vals = [float(total_reach), float(avg_reach)]
                reach_prev_vals = [float(p_total_reach), float(p_avg_reach)]
            reach_chart = _svg_bar_grouped(
                reach_labels,
                [(curr_label, reach_curr_vals, C_IG),
                 (prev_label, reach_prev_vals, '#F0B8D4')],
            )
            parts.append(
                f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px">'
                f'<div style="background:#fff;border:1px solid #f0b8d4;border-radius:10px;padding:16px">'
                f'<div style="font-size:12px;font-weight:600;color:{C_IG};margin-bottom:8px">참여 지표</div>'
                f'{eng_chart}</div>'
                f'<div style="background:#fff;border:1px solid #f0b8d4;border-radius:10px;padding:16px">'
                f'<div style="font-size:12px;font-weight:600;color:{C_IG};margin-bottom:8px">도달 지표</div>'
                f'{reach_chart}</div>'
                f'</div>'
            )
        else:
            # has_prev_ig가 False인 경우: 4개 메트릭 4-column grid
            def _ig_metric_card(label, value, sub=''):
                return (f'<div style="background:#FDF2F8;border:1px solid #F0B8D4;'
                        f'border-radius:10px;padding:14px;text-align:center">'
                        f'<div style="font-size:11px;color:{C_IG};font-weight:600;margin-bottom:6px">{label}</div>'
                        f'<div style="font-size:20px;font-weight:700;color:#333">{value}</div>'
                        f'<div style="font-size:11px;color:#888;margin-top:4px">{sub}</div>'
                        f'</div>')
            metric_cards = (
                _ig_metric_card('게시물 수', f'{post_count}개', '기간 중 업로드') +
                _ig_metric_card('총 좋아요', _f(total_likes), '기간 합산') +
                _ig_metric_card('총 도달수', _f(total_reach), '게시물 도달 합산') +
                _ig_metric_card('일평균 도달수', _f(avg_reach), '게시물 올린 날 기준')
            )
            parts.append(
                f'<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:16px">'
                f'{metric_cards}</div>'
            )

        # [C] 일별 도달수 추이 라인 차트 (이번달만, 전월 파선 없음)
        dates   = [str(d)[:10] for d in ig_account_df['날짜'].tolist()]
        reach_v = ig_account_df['도달수'].tolist()
        views_v = ig_account_df['조회수'].tolist() if '조회수' in ig_account_df.columns else [0]*len(reach_v)
        chart = _svg_dual_line(dates, reach_v, C_IG, '도달수', views_v, '#F4B400', '조회수')
        parts.append(
            f'<div style="background:#fff;border:1px solid #f0b8d4;border-radius:10px;padding:16px;margin-bottom:16px">'
            f'<div style="font-size:13px;font-weight:600;color:{C_IG};margin-bottom:8px">📈 일별 도달수 · 조회수 추이 ({curr_label})</div>'
            f'{chart}</div>'
        )

    # ── 게시물별 성과 테이블 ────────────────────────
    if ig_media_df is not None and not ig_media_df.empty:
        df = ig_media_df.copy()
        # 도달수 기준 내림차순 정렬
        if '도달수' in df.columns:
            df = df.sort_values('도달수', ascending=False)

        rows_html = ''
        type_icon = {'IMAGE': '🖼', 'VIDEO': '🎬', 'CAROUSEL_ALBUM': '📸'}
        for _, row in df.iterrows():
            날짜 = str(row.get('날짜', ''))[:10]
            유형 = row.get('유형', '')
            icon = type_icon.get(유형, '📄')
            caption = str(row.get('캡션', ''))[:30] or '(캡션 없음)'
            likes_val   = int(row.get('좋아요', 0) or 0)
            comments_val= int(row.get('댓글', 0) or 0)
            reach_val   = int(row.get('도달수', 0) or 0)
            saved   = _f(row.get('저장수', 0))
            rows_html += f'''<tr class="ig-post-row" data-reach="{reach_val}" data-likes="{likes_val}" data-comments="{comments_val}">
              <td style="padding:8px 10px;font-size:12px;color:#555">{날짜}</td>
              <td style="padding:8px 10px;font-size:12px">{icon} {유형.replace("CAROUSEL_ALBUM","카루셀").replace("IMAGE","이미지").replace("VIDEO","동영상")}</td>
              <td style="padding:8px 10px;font-size:12px;max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{caption}</td>
              <td style="padding:8px 10px;font-size:12px;text-align:right;color:{C_IG};font-weight:600">{_f(reach_val)}</td>
              <td style="padding:8px 10px;font-size:12px;text-align:right">❤️ {_f(likes_val)}</td>
              <td style="padding:8px 10px;font-size:12px;text-align:right">💬 {_f(comments_val)}</td>
              <td style="padding:8px 10px;font-size:12px;text-align:right">🔖 {saved}</td>
            </tr>'''

        parts.append(f'''
        <div style="background:#fff;border:1px solid #f0b8d4;border-radius:10px;padding:16px;overflow-x:auto">
          <div style="font-size:13px;font-weight:600;color:{C_IG};margin-bottom:12px">
            📸 게시물별 성과 (전체 {len(df)}개)
          </div>
          <div style="display:flex;align-items:center;gap:8px;margin-bottom:10px">
            <span style="font-size:12px;color:#666">정렬 기준:</span>
            <select id="ig-sort-key" onchange="igSortPosts()"
              style="font-size:12px;padding:4px 8px;border:1px solid #F0B8D4;border-radius:6px;color:#333;cursor:pointer">
              <option value="reach">도달수</option>
              <option value="likes">좋아요</option>
              <option value="comments">댓글</option>
            </select>
          </div>
          <table style="width:100%;border-collapse:collapse;font-size:13px">
            <thead>
              <tr>
                <th style="padding:8px 10px;text-align:left">날짜</th>
                <th style="padding:8px 10px;text-align:left">유형</th>
                <th style="padding:8px 10px;text-align:left">내용</th>
                <th style="padding:8px 10px;text-align:right">도달수</th>
                <th style="padding:8px 10px;text-align:right">좋아요</th>
                <th style="padding:8px 10px;text-align:right">댓글</th>
                <th style="padding:8px 10px;text-align:right">저장</th>
              </tr>
            </thead>
            <tbody id="ig-post-tbody">{rows_html}</tbody>
          </table>
        </div>
        <script>
        function igSortPosts(){{
          var key=document.getElementById("ig-sort-key").value;
          var tbody=document.getElementById("ig-post-tbody");
          var rows=Array.from(tbody.querySelectorAll("tr.ig-post-row"));
          rows.sort(function(a,b){{return parseInt(b.dataset[key]||0)-parseInt(a.dataset[key]||0);}});
          rows.forEach(function(r){{tbody.appendChild(r);}});
        }}
        </script>''')

    # ── 콘텐츠 성과 시사점 ──────────────────────
    if ig_media_df is not None and not ig_media_df.empty:
        ig_insights = _ig_insights_html(ig_media_df)
        if ig_insights:
            parts.append(ig_insights)

    if not parts:
        return '<div style="text-align:center;padding:24px;color:#888;font-size:13px">Instagram 데이터를 수집하면 이 섹션이 자동으로 표시됩니다.</div>'

    return ''.join(parts)


def _build_sns_section(sns_path: str) -> str:
    """SNS 전체 섹션 HTML — 탭 UI (전체요약 + 채널별 상세)"""
    channel_data = _read_sns_channel_data(sns_path)

    # ── 탭 바 ────────────────────────────────
    tab_bar = '<div class="sns-tab-bar">'
    tab_bar += '<button class="sns-tab sns-tab-active" onclick="snsTab(this,\'sns-p-all\')">📊 전체 요약</button>'
    for i, ch in enumerate(_SNS_CHANNEL_ORDER):
        meta  = _SNS_META[ch]
        data  = channel_data.get(ch, {})
        rows  = data.get('rows', [])
        ci1   = meta['col_idx'][1]
        filled = sum(1 for r in rows if ci1 < len(r) and r[ci1] is not None)
        badge = f'<span class="sns-tab-badge">{filled}일</span>' if filled else ''
        tab_bar += f'<button class="sns-tab" onclick="snsTab(this,\'sns-p-{i}\')">{meta["emoji"]} {ch[2:].strip()}{badge}</button>'
    tab_bar += '</div>'

    # ── 전체 요약 패널 ────────────────────────
    summary_panel = f'<div id="sns-p-all" class="sns-panel">{_sns_summary_cards(channel_data)}</div>'

    # ── 채널별 상세 패널 ──────────────────────
    channel_panels = ''
    for i, ch in enumerate(_SNS_CHANNEL_ORDER):
        meta  = _SNS_META[ch]
        color = meta['color']
        data  = channel_data.get(ch, {'rows': [], 'headers': []})
        rows  = data.get('rows', [])
        ci1   = meta['col_idx'][1]
        filled = sum(1 for r in rows if ci1 < len(r) and r[ci1] is not None)
        badge = f'<span class="sns-badge">{filled}일 입력</span>' if filled else \
                '<span class="sns-badge sns-badge-empty">미입력</span>'
        sns_block_id = f'sns{i}'
        detail = _sns_channel_detail(ch, data, block_id=sns_block_id)
        channel_panels += (
            f'<div id="sns-p-{i}" class="sns-panel" style="display:none">'
            f'<div class="sns-ch-header" style="background:{color};border-radius:8px;margin-bottom:14px">'
            f'  {ch} {badge}'
            f'</div>'
            f'<div class="sns-ch-body">{detail}</div>'
            f'</div>'
        )

    js = (
        '<script>'
        'function snsTab(btn,id){'
        'document.querySelectorAll(".sns-tab").forEach(t=>t.classList.remove("sns-tab-active"));'
        'document.querySelectorAll(".sns-panel").forEach(p=>p.style.display="none");'
        'document.getElementById(id).style.display="block";'
        'btn.classList.add("sns-tab-active");}'
        '</script>'
    )

    return tab_bar + summary_panel + channel_panels + js


# ════════════════════════════════════════════
# 메인 빌더
# ════════════════════════════════════════════

def build_html_report(raw_df, sa_conv_df=None, da_conv_df=None,
                      period_label='', output_path='report.html',
                      sns_tracker_path=None, prev_raw_df=None,
                      ig_account_df=None, ig_media_df=None,
                      prev_ig_media_df=None,
                      prev_sa_conv_df=None, prev_da_conv_df=None):

    sa_cd = _cd(sa_conv_df)
    da_cd = _cd(da_conv_df)
    # 전월 전환 상세 딕셔너리 (버튼별 MoM 비교용)
    p_sa_cd = _cd(prev_sa_conv_df) if (prev_sa_conv_df is not None and not prev_sa_conv_df.empty) else {}
    p_da_cd = _cd(prev_da_conv_df) if (prev_da_conv_df is not None and not prev_da_conv_df.empty) else {}
    today = datetime.now().strftime('%Y-%m-%d %H:%M')

    # ── 전체 집계 ──────────────────────────
    total_노출 = raw_df['노출'].sum()
    total_클릭 = raw_df['클릭'].sum()
    total_비용 = raw_df['비용'].sum()

    grp = raw_df.groupby('매체').agg(
        노출=('노출','sum'), 클릭=('클릭','sum'),
        전환=('전환','sum'), 비용=('비용','sum')
    ).reset_index()

    naver_df  = raw_df[raw_df['매체'].str.startswith('Naver')]
    gsa_df    = raw_df[raw_df['매체'] == 'Google_SA']
    gda_df    = raw_df[raw_df['매체'] == 'Google_DA']
    google_df = raw_df[raw_df['매체'].str.startswith('Google')]
    meta_df   = raw_df[raw_df['매체'] == 'Meta']
    has_meta  = not meta_df.empty

    naver_클릭  = naver_df['클릭'].sum()
    gsa_클릭    = gsa_df['클릭'].sum()
    gda_클릭    = gda_df['클릭'].sum()
    google_클릭 = google_df['클릭'].sum()
    meta_클릭   = meta_df['클릭'].sum() if has_meta else 0

    # 네이버: 별도 전환파일 없음 → raw 데이터 '총 전환수' 사용
    naver_버튼전환 = naver_df['전환'].sum()
    sa_버튼전환    = sum(sa_cd.values()) - sa_cd.get('페이지조회', 0)
    da_버튼전환    = sum(da_cd.values()) - da_cd.get('페이지조회', 0)
    meta_버튼전환  = meta_df['전환'].sum() if has_meta else 0

    # ── 매체별 집계 ───────────────────────
    naver_노출 = naver_df['노출'].sum()
    gsa_노출   = gsa_df['노출'].sum()
    gda_노출   = gda_df['노출'].sum()
    meta_노출  = meta_df['노출'].sum() if has_meta else 0
    naver_비용 = naver_df['비용'].sum()
    gsa_비용   = gsa_df['비용'].sum()
    gda_비용   = gda_df['비용'].sum()
    meta_비용  = meta_df['비용'].sum() if has_meta else 0

    # ── 전월 동기간 집계 ───────────────────
    p_노출 = p_클릭 = p_비용 = 0
    p_naver_노출 = p_gsa_노출 = p_gda_노출 = p_meta_노출 = 0
    p_naver_클릭 = p_gsa_클릭 = p_gda_클릭 = p_meta_클릭 = 0
    p_naver_비용 = p_gsa_비용 = p_gda_비용 = p_meta_비용 = 0
    has_prev = prev_raw_df is not None and not prev_raw_df.empty
    if has_prev:
        p_naver_df = prev_raw_df[prev_raw_df['매체'].str.startswith('Naver')]
        p_gsa_df   = prev_raw_df[prev_raw_df['매체'] == 'Google_SA']
        p_gda_df   = prev_raw_df[prev_raw_df['매체'] == 'Google_DA']
        p_meta_df  = prev_raw_df[prev_raw_df['매체'] == 'Meta']
        p_노출       = prev_raw_df['노출'].sum()
        p_클릭       = prev_raw_df['클릭'].sum()
        p_비용       = prev_raw_df['비용'].sum()
        p_naver_노출 = p_naver_df['노출'].sum()
        p_gsa_노출   = p_gsa_df['노출'].sum()
        p_gda_노출   = p_gda_df['노출'].sum()
        p_meta_노출  = p_meta_df['노출'].sum()
        p_naver_클릭 = p_naver_df['클릭'].sum()
        p_gsa_클릭   = p_gsa_df['클릭'].sum()
        p_gda_클릭   = p_gda_df['클릭'].sum()
        p_meta_클릭  = p_meta_df['클릭'].sum()
        p_naver_비용 = p_naver_df['비용'].sum()
        p_gsa_비용   = p_gsa_df['비용'].sum()
        p_gda_비용   = p_gda_df['비용'].sum()
        p_meta_비용  = p_meta_df['비용'].sum()

    # 기간 월 레이블 자동 도출
    def _month_label(df):
        if df is None or df.empty: return None
        try:
            return f"{pd.to_datetime(df['날짜'].min()).month}월"
        except: return None
    _curr_mon = _month_label(raw_df) or '이번달'
    _prev_mon = (_month_label(prev_raw_df) if has_prev else None) or '전월'

    # ── KPI 카드 (전체 5개 + 매체별 상세 펼침) ─
    def _mom(curr, prev): return _pct_mom(curr, prev) if has_prev else None

    # CTR / CPC 계산
    def _safe_div(a, b): return float(a) / float(b) if float(b) > 0 else 0.0
    total_CTR = _safe_div(total_클릭, total_노출) * 100
    total_CPC = _safe_div(total_비용, total_클릭)
    naver_CTR = _safe_div(naver_클릭, naver_노출) * 100
    gsa_CTR   = _safe_div(gsa_클릭,   gsa_노출)   * 100
    gda_CTR   = _safe_div(gda_클릭,   gda_노출)   * 100
    naver_CPC = _safe_div(naver_비용, naver_클릭)
    gsa_CPC   = _safe_div(gsa_비용,   gsa_클릭)
    gda_CPC   = _safe_div(gda_비용,   gda_클릭)
    meta_CTR  = _safe_div(meta_클릭,  meta_노출) * 100 if has_meta else 0.0
    meta_CPC  = _safe_div(meta_비용,  meta_클릭)       if has_meta else 0.0
    meta_전환율 = _safe_div(meta_버튼전환, meta_클릭) * 100 if has_meta else 0.0

    # 전월 CTR / CPC
    p_total_CTR = p_naver_CTR = p_gsa_CTR = p_gda_CTR = p_meta_CTR = 0.0
    p_total_CPC = p_naver_CPC = p_gsa_CPC = p_gda_CPC = p_meta_CPC = 0.0
    if has_prev:
        p_total_CTR = _safe_div(p_클릭,       p_노출)       * 100
        p_naver_CTR = _safe_div(p_naver_클릭,  p_naver_노출) * 100
        p_gsa_CTR   = _safe_div(p_gsa_클릭,    p_gsa_노출)   * 100
        p_gda_CTR   = _safe_div(p_gda_클릭,    p_gda_노출)   * 100
        p_meta_CTR  = _safe_div(p_meta_클릭,   p_meta_노출)  * 100
        p_total_CPC = _safe_div(p_비용,        p_클릭)
        p_naver_CPC = _safe_div(p_naver_비용,  p_naver_클릭)
        p_gsa_CPC   = _safe_div(p_gsa_비용,    p_gsa_클릭)
        p_gda_CPC   = _safe_div(p_gda_비용,    p_gda_클릭)
        p_meta_CPC  = _safe_div(p_meta_비용,   p_meta_클릭)

    # 버튼 전환율 계산
    total_버튼전환 = naver_버튼전환 + sa_버튼전환 + da_버튼전환
    total_전환율   = _safe_div(total_버튼전환, total_클릭) * 100
    naver_전환율   = _safe_div(naver_버튼전환, naver_클릭) * 100
    gsa_전환율     = _safe_div(sa_버튼전환,    gsa_클릭)   * 100
    gda_전환율     = _safe_div(da_버튼전환,    gda_클릭)   * 100

    # 전월 버튼 전환율
    # 전환 CSV가 있으면 _cd()로 페이지조회 제외, 없으면 raw_df '전환' 합계 사용
    p_total_전환율 = p_naver_전환율 = p_gsa_전환율 = p_gda_전환율 = p_meta_전환율 = 0.0
    p_meta_버튼전환 = 0
    if has_prev:
        p_naver_버튼전환 = p_naver_df['전환'].sum()
        # SA: 전월 전환 CSV 있으면 정확한 버튼 전환만, 없으면 raw 합계
        if prev_sa_conv_df is not None and not prev_sa_conv_df.empty:
            p_sa_cd = _cd(prev_sa_conv_df)
            p_gsa_버튼전환 = sum(p_sa_cd.values()) - p_sa_cd.get('페이지조회', 0)
        else:
            p_gsa_버튼전환 = p_gsa_df['전환'].sum()
        # DA: 전월 전환 CSV 있으면 정확한 버튼 전환만, 없으면 raw 합계
        if prev_da_conv_df is not None and not prev_da_conv_df.empty:
            p_da_cd = _cd(prev_da_conv_df)
            p_gda_버튼전환 = sum(p_da_cd.values()) - p_da_cd.get('페이지조회', 0)
        else:
            p_gda_버튼전환 = p_gda_df['전환'].sum()
        p_meta_버튼전환  = p_meta_df['전환'].sum()
        p_total_버튼전환 = p_naver_버튼전환 + p_gsa_버튼전환 + p_gda_버튼전환
        p_total_전환율   = _safe_div(p_total_버튼전환, p_클릭)      * 100
        p_naver_전환율   = _safe_div(p_naver_버튼전환, p_naver_클릭) * 100
        p_gsa_전환율     = _safe_div(p_gsa_버튼전환,   p_gsa_클릭)   * 100
        p_gda_전환율     = _safe_div(p_gda_버튼전환,   p_gda_클릭)   * 100
        p_meta_전환율    = _safe_div(p_meta_버튼전환,  p_meta_클릭)  * 100

    _bar = lambda pv, cv, fmt: (pv, cv, fmt) if has_prev else None
    kpi_html = (
        _kpi_card_expandable(
            label='전체 노출', value=_f(total_노출), sub='기간 합계',
            color=C_NAVY, mom=_mom(total_노출, p_노출),
            card_id='kpi-노출',
            bar_data=_bar(p_노출, total_노출, 'n'),
            detail_rows=[
                ('🟢 네이버',   _f(naver_노출), f'비중 {_fp(naver_노출, total_노출)}', C_NAVER,  _mom(naver_노출, p_naver_노출)),
                ('🔵 구글 SA',  _f(gsa_노출),   f'비중 {_fp(gsa_노출,   total_노출)}', C_GOOGLE, _mom(gsa_노출,   p_gsa_노출)),
                ('🔴 구글 DA',  _f(gda_노출),   f'비중 {_fp(gda_노출,   total_노출)}', C_DA,     _mom(gda_노출,   p_gda_노출)),
            ] + ([('🔷 Meta', _f(meta_노출), f'비중 {_fp(meta_노출, total_노출)}', C_META, _mom(meta_노출, p_meta_노출))] if has_meta else []),
            prev_label=_prev_mon
        ) +
        _kpi_card_expandable(
            label='전체 클릭', value=_f(total_클릭), sub=f'전체 비중 100%',
            color=C_GOOGLE, mom=_mom(total_클릭, p_클릭),
            card_id='kpi-클릭',
            bar_data=_bar(p_클릭, total_클릭, 'n'),
            detail_rows=[
                ('🟢 네이버',   _f(naver_클릭), f'비중 {_fp(naver_클릭, total_클릭)}', C_NAVER,  _mom(naver_클릭, p_naver_클릭)),
                ('🔵 구글 SA',  _f(gsa_클릭),   f'비중 {_fp(gsa_클릭,   total_클릭)}', C_GOOGLE, _mom(gsa_클릭,   p_gsa_클릭)),
                ('🔴 구글 DA',  _f(gda_클릭),   f'비중 {_fp(gda_클릭,   total_클릭)}', C_DA,     _mom(gda_클릭,   p_gda_클릭)),
            ] + ([('🔷 Meta', _f(meta_클릭), f'비중 {_fp(meta_클릭, total_클릭)}', C_META, _mom(meta_클릭, p_meta_클릭))] if has_meta else []),
            prev_label=_prev_mon
        ) +
        _kpi_card_expandable(
            label='전체 광고비', value=f'₩{_f(total_비용)}',
            sub='기간 합계',
            color=C_DA, mom=_mom(total_비용, p_비용),
            card_id='kpi-비용',
            bar_data=_bar(p_비용, total_비용, 'won'),
            detail_rows=[
                ('🟢 네이버',   f'₩{_f(naver_비용)}', f'비중 {_fp(naver_비용, total_비용)}', C_NAVER,  _mom(naver_비용, p_naver_비용)),
                ('🔵 구글 SA',  f'₩{_f(gsa_비용)}',   f'비중 {_fp(gsa_비용,   total_비용)}', C_GOOGLE, _mom(gsa_비용,   p_gsa_비용)),
                ('🔴 구글 DA',  f'₩{_f(gda_비용)}',   f'비중 {_fp(gda_비용,   total_비용)}', C_DA,     _mom(gda_비용,   p_gda_비용)),
            ] + ([('🔷 Meta', f'₩{_f(meta_비용)}', f'비중 {_fp(meta_비용, total_비용)}', C_META, _mom(meta_비용, p_meta_비용))] if has_meta else []),
            prev_label=_prev_mon
        ) +
        _kpi_card_expandable(
            label='전체 CTR', value=f'{total_CTR:.2f}%',
            sub='클릭률 (클릭 ÷ 노출)',
            color=C_GOLD, mom=_mom(total_CTR, p_total_CTR),
            card_id='kpi-ctr',
            bar_data=_bar(p_total_CTR, total_CTR, 'pct'),
            detail_rows=[
                ('🟢 네이버',   f'{naver_CTR:.2f}%', f'클릭 {_f(naver_클릭)}', C_NAVER,  _mom(naver_CTR, p_naver_CTR)),
                ('🔵 구글 SA',  f'{gsa_CTR:.2f}%',   f'클릭 {_f(gsa_클릭)}',   C_GOOGLE, _mom(gsa_CTR,   p_gsa_CTR)),
                ('🔴 구글 DA',  f'{gda_CTR:.2f}%',   f'클릭 {_f(gda_클릭)}',   C_DA,     _mom(gda_CTR,   p_gda_CTR)),
            ] + ([('🔷 Meta', f'{meta_CTR:.2f}%', f'클릭 {_f(meta_클릭)}', C_META, _mom(meta_CTR, p_meta_CTR))] if has_meta else []),
            prev_label=_prev_mon
        ) +
        _kpi_card_expandable(
            label='평균 CPC', value=f'₩{_f(total_CPC)}',
            sub='클릭당 광고비',
            color=C_MUTED, mom=_mom(total_CPC, p_total_CPC),
            card_id='kpi-cpc',
            bar_data=_bar(p_total_CPC, total_CPC, 'won'),
            detail_rows=[
                ('🟢 네이버',   f'₩{_f(naver_CPC)}', f'비용 {_f(naver_비용)}', C_NAVER,  _mom(naver_CPC, p_naver_CPC)),
                ('🔵 구글 SA',  f'₩{_f(gsa_CPC)}',   f'비용 {_f(gsa_비용)}',   C_GOOGLE, _mom(gsa_CPC,   p_gsa_CPC)),
                ('🔴 구글 DA',  f'₩{_f(gda_CPC)}',   f'비용 {_f(gda_비용)}',   C_DA,     _mom(gda_CPC,   p_gda_CPC)),
            ] + ([('🔷 Meta', f'₩{_f(meta_CPC)}', f'비용 {_f(meta_비용)}', C_META, _mom(meta_CPC, p_meta_CPC))] if has_meta else []),
            prev_label=_prev_mon
        ) +
        _kpi_card_expandable(
            label='버튼 전환율', value=f'{total_전환율:.2f}%',
            sub=f'총 {_f(total_버튼전환)}건 전환',
            color='#7C3AED', mom=_mom(total_전환율, p_total_전환율),
            card_id='kpi-conv',
            bar_data=_bar(p_total_전환율, total_전환율, 'pct'),
            detail_rows=[
                ('🟢 네이버',   f'{_f(naver_버튼전환)}건', f'전환율 {naver_전환율:.2f}%', C_NAVER,  _mom(naver_전환율, p_naver_전환율)),
                ('🔵 구글 SA',  f'{_f(sa_버튼전환)}건',    f'전환율 {gsa_전환율:.2f}%',   C_GOOGLE, _mom(gsa_전환율,   p_gsa_전환율)),
                ('🔴 구글 DA',  f'{_f(da_버튼전환)}건',    f'전환율 {gda_전환율:.2f}%',   C_DA,     _mom(gda_전환율,   p_gda_전환율)),
            ] + ([('🔷 Meta', f'{_f(meta_버튼전환)}건', f'전환율 {meta_전환율:.2f}%', C_META, _mom(meta_전환율, p_meta_전환율))] if has_meta else []),
            prev_label=_prev_mon
        )
    )
    kpi_section = f'<div class="kpi-grid-6">{kpi_html}</div>'

    # ── 데이터 기반 시사점 ──────────────────────
    insights_content = _insights_html(
        total_노출=total_노출, total_클릭=total_클릭, total_비용=total_비용,
        total_CTR=total_CTR, total_CPC=total_CPC,
        naver_클릭=naver_클릭, naver_비용=naver_비용, naver_CTR=naver_CTR, naver_CPC=naver_CPC,
        gsa_클릭=gsa_클릭,     gsa_비용=gsa_비용,     gsa_CTR=gsa_CTR,     gsa_CPC=gsa_CPC,
        gda_클릭=gda_클릭,     gda_비용=gda_비용,     gda_CTR=gda_CTR,     gda_CPC=gda_CPC,
        has_prev=has_prev,
        p_노출=p_노출, p_클릭=p_클릭, p_비용=p_비용,
        p_total_CTR=p_total_CTR, p_total_CPC=p_total_CPC,
    )
    insights_section = _section('📌 데이터 기반 시사점', C_NAVY, insights_content) if insights_content else ''

    # 매체별 버튼 전환 매핑
    _btn_map = {}
    for _, r in grp.iterrows():
        m = str(r['매체'])
        if m.startswith('Naver'):
            _btn_map[m] = naver_버튼전환
        elif m == 'Google_SA':
            _btn_map[m] = sa_버튼전환
        elif m == 'Google_DA':
            _btn_map[m] = da_버튼전환
        else:
            _btn_map[m] = 0

    # ── 플랫폼 비교 테이블 ─────────────────
    comp_rows = []
    for _, r in grp.iterrows():
        노출 = float(r['노출']); 클릭 = float(r['클릭']); 비용 = float(r['비용'])
        btn  = _btn_map.get(str(r['매체']), 0)
        comp_rows.append([
            r['매체'],
            _f(노출), _f(클릭), _fp(클릭, 노출),
            f'₩{_f(비용/클릭 if 클릭 else 0)}', f'₩{_f(비용)}',
            _fp(노출, total_노출), _fp(클릭, total_클릭), _fp(비용, total_비용),
            f'{_f(btn)}건 ({_fp(btn, 클릭)})',
        ])
    comp_tbl = _table_html(
        ['매체','노출','클릭','CTR','CPC','비용','노출비중','클릭비중','비용비중','버튼 전환'],
        comp_rows
    )

    # ── 플랫폼 비교 차트 ───────────────────
    media_labels = list(grp['매체'])
    click_vals   = [float(v) for v in grp['클릭']]
    cost_vals    = [float(v) for v in grp['비용']]
    media_colors = [MEDIA_COLORS.get(m.split('_')[0] if '_' in m else m, C_NAVY) for m in media_labels]

    svg_cmp_click = _svg_bar_grouped(
        media_labels,
        [('클릭수', click_vals, C_GOOGLE)],
    )
    svg_cost_pie = _svg_pie(
        media_labels, cost_vals, media_colors,
    )

    compare_content = (
        f'<div class="two-col">'
        f'<div class="col-main">{comp_tbl}</div>'
        f'</div>'
        f'<div class="chart-row">'
        f'<div class="chart-item">{_chart_block("매체별 클릭수 비교", svg_cmp_click)}</div>'
        f'<div class="chart-item chart-narrow">{_chart_block("매체별 비용 비중", svg_cost_pie)}</div>'
        f'</div>'
    )

    # ── 매체별 일별/주별 트렌드 ───────────────
    def _daily_block(df, media_label, color, block_id='block'):
        if df.empty: return ''

        # ── 일별 집계 ──
        daily = df.groupby('날짜', as_index=False).agg(
            노출=('노출','sum'), 클릭=('클릭','sum'), 비용=('비용','sum')
        ).sort_values('날짜')
        daily['날짜_str'] = pd.to_datetime(daily['날짜']).dt.strftime('%Y-%m-%d')
        dates  = list(daily['날짜_str'])
        clicks = [float(v) for v in daily['클릭']]
        costs  = [float(v) for v in daily['비용']]

        svg_c_d = _svg_line(dates, clicks, color)
        svg_v_d = _svg_line(dates, costs,  color)

        # 일별 테이블 (data-date 속성 포함 → JS 기간 필터용)
        daily_rows_html = ''
        for i, r in daily.iterrows():
            클릭 = float(r['클릭']); 노출 = float(r['노출']); 비용 = float(r['비용'])
            stripe = ' class="stripe"' if i % 2 == 0 else ''
            daily_rows_html += (
                f'<tr class="daily-row"{stripe} data-date="{r["날짜_str"]}">'
                f'<td>{r["날짜_str"]}</td>'
                f'<td>{_f(노출)}</td><td>{_f(클릭)}</td>'
                f'<td class="pct">{_fp(클릭, 노출)}</td>'
                f'<td>₩{_f(비용/클릭 if 클릭 else 0)}</td>'
                f'<td>₩{_f(비용)}</td>'
                f'</tr>'
            )
        daily_tbl = (
            f'<div style="overflow-x:auto">'
            f'<table><thead><tr>'
            f'<th>날짜</th><th>노출</th><th>클릭</th><th>CTR</th><th>CPC</th><th>비용</th>'
            f'</tr></thead><tbody>{daily_rows_html}</tbody></table></div>'
        )

        # ── 주별 집계 ──
        df2 = df.copy()
        df2['날짜'] = pd.to_datetime(df2['날짜'])
        df2['주'] = df2['날짜'].dt.to_period('W')
        weekly = df2.groupby('주', as_index=False).agg(
            노출=('노출','sum'), 클릭=('클릭','sum'), 비용=('비용','sum')
        ).sort_values('주')
        weekly['주_str'] = weekly['주'].apply(
            lambda p: f"{p.start_time.strftime('%m/%d')}~{p.end_time.strftime('%m/%d')}"
        )
        w_dates  = list(weekly['주_str'])
        w_clicks = [float(v) for v in weekly['클릭']]
        w_costs  = [float(v) for v in weekly['비용']]

        svg_c_w = _svg_line(w_dates, w_clicks, color)
        svg_v_w = _svg_line(w_dates, w_costs,  color)

        weekly_rows_html = ''
        for i, r in weekly.iterrows():
            클릭 = float(r['클릭']); 노출 = float(r['노출']); 비용 = float(r['비용'])
            stripe = ' class="stripe"' if i % 2 == 0 else ''
            weekly_rows_html += (
                f'<tr{stripe}>'
                f'<td>{r["주_str"]}</td>'
                f'<td>{_f(노출)}</td><td>{_f(클릭)}</td>'
                f'<td class="pct">{_fp(클릭, 노출)}</td>'
                f'<td>₩{_f(비용/클릭 if 클릭 else 0)}</td>'
                f'<td>₩{_f(비용)}</td>'
                f'</tr>'
            )
        weekly_tbl = (
            f'<div style="overflow-x:auto">'
            f'<table><thead><tr>'
            f'<th>주간</th><th>노출</th><th>클릭</th><th>CTR</th><th>CPC</th><th>비용</th>'
            f'</tr></thead><tbody>{weekly_rows_html}</tbody></table></div>'
        )

        return f'''
<div class="aggr-toggle">
  <button class="aggr-btn aggr-active" id="aggr-daily-{block_id}"
    onclick="setAggr('{block_id}','daily')">일별</button>
  <button class="aggr-btn" id="aggr-weekly-{block_id}"
    onclick="setAggr('{block_id}','weekly')">주별</button>
</div>
<div id="daily-view-{block_id}">
  <div class="chart-row">
    <div class="chart-item">{_chart_block("일별 클릭수 추이", svg_c_d)}</div>
    <div class="chart-item">{_chart_block("일별 비용 추이", svg_v_d)}</div>
  </div>
  <div class="sub-title">일별 성과 상세</div>
  {daily_tbl}
</div>
<div id="weekly-view-{block_id}" style="display:none">
  <div class="chart-row">
    <div class="chart-item">{_chart_block("주별 클릭수 추이", svg_c_w)}</div>
    <div class="chart-item">{_chart_block("주별 비용 추이", svg_v_w)}</div>
  </div>
  <div class="sub-title">주별 성과 상세</div>
  {weekly_tbl}
</div>
'''

    naver_content  = _daily_block(naver_df,  '네이버 SA',  C_NAVER,  'naver')
    gsa_content    = _daily_block(gsa_df,    '구글 SA',    C_GOOGLE, 'gsa')
    gda_content    = _daily_block(gda_df,    '구글 DA',    C_DA,     'gda')

    # 네이버 버튼 전환 블록 (총 전환수 기준)
    naver_conv_html = (
        f'<div class="sub-title">네이버 SA 버튼 전환 (총 전환수 기준)</div>'
        + _table_html(
            ['전환 기준', '전환수', '전환율(클릭 대비)'],
            [['총 전환수 (상담예약·신청·상세보기 통합)',
              f'{_f(naver_버튼전환)}건',
              _fp(naver_버튼전환, naver_클릭)]],
            stripe=False
        )
        + '<p style="font-size:11px;color:#718096;margin-top:6px">'
          '※ 네이버는 버튼별 세분화 미설정으로 총 전환수를 버튼 전환으로 사용합니다.</p>'
    )

    # ── 버튼 전환 섹션 ─────────────────────
    def _conv_block(cd, 클릭수, label, color, prev_cd=None, prev_클릭수=0,
                    curr_label='이번달', prev_label='전월'):
        if not cd: return ''
        상세 = cd.get('상세보기_버튼클릭', 0)
        예약 = cd.get('상담예약_버튼클릭', 0)
        신청 = cd.get('상담신청_버튼클릭', 0)
        합계 = 상세 + 예약 + 신청

        if prev_cd:
            p상세 = prev_cd.get('상세보기_버튼클릭', 0)
            p예약 = prev_cd.get('상담예약_버튼클릭', 0)
            p신청 = prev_cd.get('상담신청_버튼클릭', 0)
            p합계 = p상세 + p예약 + p신청
            def _d(cur, prv):
                if prv == 0: return '-'
                delta = int(round(cur - prv))
                sign, clr = ('▲', '#2F855A') if delta >= 0 else ('▼', '#C53030')
                return f'<span style="color:{clr};font-weight:bold">{sign}{abs(delta):,}</span>'
            rows = [
                ['상세보기', f'{_f(상세)}건', _fp(상세, 클릭수),  f'{_f(p상세)}건', _fp(p상세, prev_클릭수), _d(상세, p상세)],
                ['상담예약', f'{_f(예약)}건', _fp(예약, 클릭수),  f'{_f(p예약)}건', _fp(p예약, prev_클릭수), _d(예약, p예약)],
                ['상담신청', f'{_f(신청)}건', _fp(신청, 클릭수),  f'{_f(p신청)}건', _fp(p신청, prev_클릭수), _d(신청, p신청)],
                ['합계',     f'{_f(합계)}건', _fp(합계, 클릭수),  f'{_f(p합계)}건', _fp(p합계, prev_클릭수), _d(합계, p합계)],
            ]
            tbl = _table_html(['전환 유형', curr_label, f'{curr_label} 전환율', prev_label, f'{prev_label} 전환율', '증감'], rows)
        else:
            rows = [
                ['상세보기 버튼클릭', _f(상세), _fp(상세, 클릭수)],
                ['상담예약 버튼클릭', _f(예약), _fp(예약, 클릭수)],
                ['상담신청 버튼클릭', _f(신청), _fp(신청, 클릭수)],
                ['버튼클릭 합계',     _f(합계), _fp(합계, 클릭수)],
            ]
            tbl = _table_html(['전환 유형', '전환수', f'전환율({label} 클릭 대비)'], rows)

        return f'<div class="sub-title">{label} 버튼 전환</div>{tbl}'

    sa_conv_content = _conv_block(sa_cd, gsa_클릭, 'SA', C_GOOGLE, prev_cd=p_sa_cd, prev_클릭수=p_gsa_클릭,
                                   curr_label=_curr_mon, prev_label=_prev_mon)
    da_conv_content = _conv_block(da_cd, gda_클릭, 'DA', C_DA,     prev_cd=p_da_cd, prev_클릭수=p_gda_클릭,
                                   curr_label=_curr_mon, prev_label=_prev_mon)

    # SA vs DA 비교 바차트 (전월 데이터 있으면 이번달 vs 전월 함께 표시)
    btn_labels = ['상세보기', '상담예약', '상담신청']
    sa_vals = [sa_cd.get('상세보기_버튼클릭',0), sa_cd.get('상담예약_버튼클릭',0), sa_cd.get('상담신청_버튼클릭',0)]
    da_vals = [da_cd.get('상세보기_버튼클릭',0), da_cd.get('상담예약_버튼클릭',0), da_cd.get('상담신청_버튼클릭',0)]
    if p_sa_cd or p_da_cd:
        p_sa_vals = [p_sa_cd.get('상세보기_버튼클릭',0), p_sa_cd.get('상담예약_버튼클릭',0), p_sa_cd.get('상담신청_버튼클릭',0)]
        p_da_vals = [p_da_cd.get('상세보기_버튼클릭',0), p_da_cd.get('상담예약_버튼클릭',0), p_da_cd.get('상담신청_버튼클릭',0)]
        svg_conv_bar = _svg_bar_grouped(
            btn_labels,
            [(f'SA {_curr_mon}', sa_vals, C_GOOGLE), (f'SA {_prev_mon}', p_sa_vals, '#A8C4F5'),
             (f'DA {_curr_mon}', da_vals, C_DA),     (f'DA {_prev_mon}', p_da_vals, '#F2A9A2')],
        )
        conv_chart_title = f'SA · DA 버튼 전환 — {_curr_mon} vs {_prev_mon}'
    else:
        svg_conv_bar = _svg_bar_grouped(
            btn_labels,
            [('Google SA', sa_vals, C_GOOGLE), ('Google DA', da_vals, C_DA)],
        )
        conv_chart_title = 'SA vs DA 버튼별 전환 비교'

    conv_content = (
        f'{_chart_block(conv_chart_title, svg_conv_bar)}'
        f'<div class="two-col-conv">'
        f'<div>{sa_conv_content}</div>'
        f'<div>{da_conv_content}</div>'
        f'</div>'
    )

    # ══════════════════════════════════════
    # CSS
    # ══════════════════════════════════════
    css = f"""
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{
    font-family: 'Apple SD Gothic Neo','Malgun Gothic','맑은 고딕',Arial,sans-serif;
    background: {C_LIGHT}; color: {C_TEXT}; font-size: 13px;
}}
.page {{ max-width: 1200px; margin: 0 auto; padding: 24px 16px; }}

/* 헤더 */
.header {{ background: {C_NAVY}; color: white; border-radius: 10px;
    padding: 20px 28px; margin-bottom: 20px; }}
.header h1 {{ font-size: 20px; font-weight: 700; }}
.header .meta {{ font-size: 12px; opacity: 0.7; margin-top: 6px; }}

/* KPI 카드 — 전체 6개 (3×2) + 상세 펼침 */
.kpi-grid-6 {{ display: grid; grid-template-columns: repeat(3,1fr); gap: 14px; margin-bottom: 20px; }}
.kpi-expand-wrap {{ display: flex; flex-direction: column; }}
.kpi-card {{ background: white; border-radius: 8px; padding: 14px 16px;
    box-shadow: 0 1px 4px rgba(0,0,0,.08); }}
.kpi-label {{ font-size: 11px; color: {C_MUTED}; font-weight: 500; margin-bottom: 6px; }}
.kpi-value {{ font-size: 20px; font-weight: 700; }}
.kpi-sub {{ font-size: 11px; color: {C_MUTED}; margin-top: 4px; }}
.kpi-mom {{ font-size: 11px; font-weight: 700; margin-top: 6px; padding: 3px 8px;
    background: #F8FAFC; border-radius: 12px; display: inline-block; }}
.kpi-detail-btn {{ margin-top: 10px; width: 100%; padding: 5px 0;
    background: {C_LIGHT}; border: none; border-radius: 6px;
    font-size: 11px; font-weight: 600; color: {C_MUTED}; cursor: pointer;
    transition: background .15s; font-family: inherit; }}
.kpi-detail-btn:hover {{ background: {C_BORDER}; color: {C_TEXT}; }}
.kpi-detail-panel {{ display: none; background: white;
    border: 1px solid {C_BORDER}; border-top: none;
    border-radius: 0 0 8px 8px; overflow: hidden;
    box-shadow: 0 2px 6px rgba(0,0,0,.06); }}
.kpi-detail-row {{ display: flex; align-items: center; gap: 8px;
    padding: 9px 14px; border-bottom: 1px solid {C_BORDER}; flex-wrap: wrap; }}
.kpi-detail-row:last-child {{ border-bottom: none; }}
.kpi-dr-label {{ font-size: 12px; font-weight: 700; min-width: 72px; }}
.kpi-dr-val {{ font-size: 14px; font-weight: 700; color: {C_TEXT}; flex: 1; }}
.kpi-dr-sub {{ font-size: 11px; color: {C_MUTED}; }}
.kpi-dr-mom {{ font-size: 11px; font-weight: 700; padding: 2px 7px;
    background: #F8FAFC; border-radius: 10px; margin-left: auto; }}

/* 섹션 */
.section {{ background: white; border-radius: 8px; box-shadow: 0 1px 4px rgba(0,0,0,.08);
    margin-bottom: 20px; overflow: hidden; }}
.section-header {{ color: white; font-weight: 700; font-size: 14px;
    padding: 12px 20px; }}
.section-body {{ padding: 20px; }}

/* 테이블 */
table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
thead tr {{ background: {C_NAVY}; color: white; }}
th {{ padding: 9px 10px; text-align: center; font-weight: 600; white-space: nowrap; }}
td {{ padding: 8px 10px; text-align: center; border-bottom: 1px solid {C_BORDER}; }}
tr.stripe {{ background: #F8FAFC; }}
td.pct {{ background: #FFFBEB; font-weight: 600; color: #92400E; }}
.sub-title {{ font-size: 13px; font-weight: 700; color: {C_NAVY};
    margin: 16px 0 8px; padding-left: 10px; border-left: 4px solid {C_NAVY}; }}

/* 차트 */
.chart-row {{ display: flex; gap: 16px; margin: 16px 0; }}
.chart-item {{ flex: 1; min-width: 0; background: white;
    border: 1px solid {C_BORDER}; border-radius: 8px; overflow: hidden; }}
.chart-narrow {{ max-width: 380px; flex: 0 0 380px; }}
.chart-block {{ width: 100%; padding: 12px; }}
.chart-title {{ font-size: 12px; font-weight: 600; color: {C_MUTED};
    text-align: center; margin-bottom: 8px; }}
.two-col-conv {{ display: grid; grid-template-columns: 1fr 1fr; gap: 24px; margin-top: 16px; }}

/* 시사점 섹션 */
.ins-grid {{ display: grid; grid-template-columns: repeat(3,1fr); gap: 16px; }}
.ins-block {{ border-radius: 10px; padding: 16px 18px; }}
.ins-good  {{ background: #F0FDF4; border: 1px solid #86EFAC; }}
.ins-warn  {{ background: #FFFBEB; border: 1px solid #FCD34D; }}
.ins-rec   {{ background: #EFF6FF; border: 1px solid #93C5FD; }}
.ins-title {{ font-size: 13px; font-weight: 700; margin-bottom: 10px; color: {C_TEXT}; }}
.ins-block ul {{ margin: 0; padding-left: 18px; }}
.ins-block li {{ font-size: 13px; color: #374151; line-height: 1.65; margin-bottom: 7px; }}
.ins-block li:last-child {{ margin-bottom: 0; }}

/* SNS 섹션 */
.sns-kpi-grid {{ display:grid; grid-template-columns:repeat(5,1fr); gap:12px; margin-bottom:20px; }}
.sns-kpi {{ background:white; border-radius:8px; padding:14px 16px;
    box-shadow:0 1px 4px rgba(0,0,0,.08); }}
.sns-ch-name {{ font-size:12px; font-weight:700; margin-bottom:4px; }}
.sns-metric-label {{ font-size:11px; color:{C_MUTED}; margin-bottom:6px; }}
.sns-value {{ font-size:22px; font-weight:800; margin-bottom:4px; }}
.sns-delta {{ font-size:11px; font-weight:600; }}
.sns-ch-header {{ color:white; font-weight:700; font-size:13px;
    padding:10px 16px; display:flex; align-items:center; justify-content:space-between; }}
.sns-ch-body {{ padding:4px 0; }}
.sns-badge {{ background:rgba(255,255,255,0.25); border-radius:10px;
    padding:2px 10px; font-size:11px; font-weight:600; }}
.sns-badge-empty {{ background:rgba(0,0,0,0.20); }}
/* SNS 탭 */
.sns-tab-bar {{ display:flex; gap:4px; flex-wrap:wrap; margin-bottom:16px;
    background:white; padding:6px; border-radius:10px;
    box-shadow:0 1px 4px rgba(0,0,0,.08); }}
.sns-tab {{ flex:1; min-width:80px; padding:8px 6px; text-align:center;
    border-radius:7px; cursor:pointer; font-weight:600; font-size:12px;
    border:none; background:transparent; color:{C_MUTED}; transition:all .15s; }}
.sns-tab-active {{ background:{C_NAVY}; color:white; }}
.sns-tab-badge {{ background:rgba(255,255,255,0.3); border-radius:8px;
    padding:1px 6px; font-size:10px; margin-left:4px; }}
.sns-panel {{ }}
/* SNS 통계 요약 스트립 */
.stats-strip {{ display:flex; flex-wrap:wrap; gap:10px; margin-bottom:14px; }}
.stat-box {{ background:{C_LIGHT}; border-radius:8px; padding:10px 14px; min-width:100px; flex:1; }}
.stat-label {{ font-size:10px; color:{C_MUTED}; font-weight:600; margin-bottom:4px; }}
.stat-val {{ font-size:15px; font-weight:800; color:{C_TEXT}; }}
.stat-sub {{ font-size:10px; color:{C_MUTED}; margin-top:2px; }}

/* 기간 필터 바 */
.period-bar {{ display:flex; align-items:center; gap:6px; flex-wrap:wrap;
    background:white; padding:10px 16px; border-radius:8px;
    box-shadow:0 1px 4px rgba(0,0,0,.08); margin-bottom:4px; }}
.period-label {{ font-size:12px; font-weight:600; color:{C_MUTED}; margin-right:4px; }}
.period-btn {{ padding:5px 16px; border-radius:15px;
    border:1.5px solid {C_BORDER}; background:white;
    color:{C_MUTED}; font-size:12px; font-weight:600; cursor:pointer;
    transition:all .15s; font-family:inherit; }}
.period-btn:hover {{ border-color:{C_NAVY}; color:{C_NAVY}; }}
.period-active {{ background:{C_NAVY} !important; color:white !important;
    border-color:{C_NAVY} !important; }}
/* 집계 토글 */
.aggr-toggle {{ display:flex; gap:6px; margin-bottom:10px; }}
.aggr-btn {{ padding:5px 16px; border-radius:15px;
    border:1.5px solid {C_BORDER}; background:white;
    color:{C_MUTED}; font-size:12px; font-weight:600; cursor:pointer;
    transition:all .15s; font-family:inherit; }}
.aggr-btn:hover {{ border-color:{C_NAVY}; color:{C_NAVY}; }}
.aggr-active {{ background:{C_NAVY} !important; color:white !important;
    border-color:{C_NAVY} !important; }}

/* 메인 탭 네비게이션 */
.main-tab-bar {{
  display: flex; gap: 6px;
  background: white; padding: 6px;
  border-radius: 14px;
  box-shadow: 0 2px 10px rgba(0,0,0,.10);
  margin-bottom: 20px;
  position: sticky; top: 10px; z-index: 200;
}}
.main-tab {{
  flex: 1; padding: 11px 10px; text-align: center;
  border-radius: 10px; cursor: pointer;
  font-weight: 700; font-size: 14px;
  border: none; background: transparent; color: #777;
  transition: all .18s; font-family: inherit; line-height: 1.3;
}}
.main-tab:hover {{ background: #f0f4fa; color: #333; }}
.main-tab-active {{ background: {C_NAVY}; color: white; box-shadow: 0 2px 8px rgba(27,58,107,.25); }}
.tab-panel {{ display: none; }}
.tab-panel-active {{ display: block; }}

/* 반응형 */
@media (max-width: 900px) {{
    .kpi-grid-6 {{ grid-template-columns: repeat(2,1fr); }}
    .ins-grid {{ grid-template-columns: 1fr; }}
    .chart-row {{ flex-direction: column; }}
    .chart-narrow {{ max-width: 100%; flex: 1; }}
    .two-col-conv {{ grid-template-columns: 1fr; }}
    .sns-kpi-grid {{ grid-template-columns: repeat(2,1fr); }}
    .sns-grid {{ grid-template-columns: 1fr; }}
    .main-tab {{ font-size: 12px; padding: 9px 6px; }}
}}
"""

    # ══════════════════════════════════════
    # HTML 조립
    # ══════════════════════════════════════
    html_body = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>광고 성과 대시보드 | {period_label}</title>
<style>{css}</style>
</head>
<body>
<div class="page">

  <div class="header">
    <h1>📊 광고 성과 대시보드</h1>
    <div class="meta">기간: {period_label} &nbsp;|&nbsp; 생성: {today}</div>
  </div>

  {kpi_section}

  {insights_section}

  <!-- ── 메인 탭 바 ── -->
  <div class="main-tab-bar">
    <button class="main-tab main-tab-active" id="tab-btn-ad"  onclick="switchTab('ad')">📊 광고 성과</button>
    <button class="main-tab" id="tab-btn-ig"  onclick="switchTab('ig')">📸 Instagram 유기</button>
    <button class="main-tab" id="tab-btn-sns" onclick="switchTab('sns')">📱 SNS 채널</button>
  </div>

  <!-- ── 광고 성과 탭 ── -->
  <div class="tab-panel tab-panel-active" id="tab-panel-ad">

    {_section('🔍 전체 매체 비교', C_NAVY, compare_content)}

    <div class="period-bar">
      <span class="period-label">📅 기간 필터</span>
      <button class="period-btn" onclick="setPeriod(this,7)">최근 7일</button>
      <button class="period-btn" onclick="setPeriod(this,14)">최근 14일</button>
      <button class="period-btn period-active" onclick="setPeriod(this,0)">전체</button>
    </div>

    {_section('🟢 네이버 SA', C_NAVER, naver_content + naver_conv_html)}

    {_section('🔵 구글 SA (검색)', C_GOOGLE, gsa_content + sa_conv_content)}

    {_section('🔴 구글 DA (실적 최대화)', C_DA, gda_content + da_conv_content)}

    {_section('🎯 버튼 전환 상세 (SA / DA 비교)', C_GOOGLE, conv_content)}

  </div>

  <!-- ── Instagram 유기 탭 ── -->
  <div class="tab-panel" id="tab-panel-ig">

    {_section('📸 Instagram 유기 성과 (@ktplaza_story)', C_IG,
        _build_instagram_section(ig_account_df, ig_media_df, prev_ig_media_df)
    )}

  </div>

  <!-- ── SNS 채널 탭 ── -->
  <div class="tab-panel" id="tab-panel-sns">

    {_section('📱 SNS 채널 현황', '#6C3483',
        _build_sns_section(sns_tracker_path) if sns_tracker_path else
        '<div style="text-align:center;padding:24px;color:#888;font-size:13px">'
        '광고 보고서 생성 시 <b>SNS_채널_관리대장.xlsx</b>를 함께 제공하면 이 섹션이 자동으로 표시됩니다.</div>'
    )}

  </div>

</div>

<!-- Instagram 호버 툴팁 -->
<div id="ig-tooltip" style="position:fixed;display:none;background:rgba(30,58,95,0.93);
  color:white;padding:9px 14px;border-radius:10px;font-size:12px;line-height:1.7;
  pointer-events:none;z-index:9999;box-shadow:0 4px 16px rgba(0,0,0,.25);
  white-space:nowrap"></div>

<script>
/* ── Instagram 차트 툴팁 ── */
function showIgTip(e, date, reach, views) {{
  var tip = document.getElementById('ig-tooltip');
  tip.innerHTML = '<b style="font-size:13px">' + date + '</b><br>'
    + '<span style="color:#F9A8C9">● 도달수</span>  <b>' + reach + '</b><br>'
    + '<span style="color:#FCD34D">● 조회수</span>  <b>' + views + '</b>';
  tip.style.display = 'block';
  tip.style.left = (e.clientX + 16) + 'px';
  tip.style.top  = (e.clientY - 10) + 'px';
}}
function hideIgTip() {{
  document.getElementById('ig-tooltip').style.display = 'none';
}}

/* ── 메인 탭 전환 ── */
function switchTab(tabId) {{
  document.querySelectorAll('.main-tab').forEach(function(t) {{
    t.classList.remove('main-tab-active');
  }});
  document.querySelectorAll('.tab-panel').forEach(function(p) {{
    p.classList.remove('tab-panel-active');
  }});
  document.getElementById('tab-btn-' + tabId).classList.add('main-tab-active');
  document.getElementById('tab-panel-' + tabId).classList.add('tab-panel-active');
  window.scrollTo({{top: 0, behavior: 'smooth'}});
}}

/* ── KPI 상세 펼침 토글 ── */
function toggleKpi(id) {{
  var panel = document.getElementById(id);
  var icon  = document.getElementById(id + '-icon');
  var open  = panel.style.display === 'block';
  panel.style.display = open ? 'none' : 'block';
  if (icon) icon.textContent = open ? '▼' : '▲';
}}

/* ── 집계 토글: 일별 ↔ 주별 ── */
function setAggr(blockId, aggr) {{
  document.getElementById('daily-view-'  + blockId).style.display = aggr === 'daily'  ? '' : 'none';
  document.getElementById('weekly-view-' + blockId).style.display = aggr === 'weekly' ? '' : 'none';
  document.getElementById('aggr-daily-'  + blockId).className = 'aggr-btn' + (aggr === 'daily'  ? ' aggr-active' : '');
  document.getElementById('aggr-weekly-' + blockId).className = 'aggr-btn' + (aggr === 'weekly' ? ' aggr-active' : '');
  /* 기간 필터 재적용 */
  applyPeriod(_activeDays);
}}

/* ── 기간 필터: 최근 N일 (0 = 전체) ── */
var _activeDays = 0;
function setPeriod(btn, days) {{
  _activeDays = days;
  document.querySelectorAll('.period-btn').forEach(function(b) {{
    b.classList.remove('period-active');
  }});
  btn.classList.add('period-active');
  applyPeriod(days);
}}

function applyPeriod(days) {{
  var rows = document.querySelectorAll('.daily-row');
  if (days === 0) {{
    rows.forEach(function(r) {{ r.style.display = ''; }});
    return;
  }}
  /* 전체 날짜 목록에서 최신 N개의 cutoff 날짜 계산 */
  var allDates = [];
  rows.forEach(function(r) {{
    var d = r.getAttribute('data-date');
    if (d && allDates.indexOf(d) === -1) allDates.push(d);
  }});
  allDates.sort();
  var cutoff = allDates.length > days ? allDates[allDates.length - days] : (allDates[0] || '');
  rows.forEach(function(r) {{
    r.style.display = (r.getAttribute('data-date') >= cutoff) ? '' : 'none';
  }});
}}
</script>
</body>
</html>"""

    Path(output_path).write_text(html_body, encoding='utf-8')
    print(f'✅ HTML 대시보드 저장: {output_path}')
    return output_path
