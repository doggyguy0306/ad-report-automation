"""
광고 Excel 장표 생성 모듈
시트 구성:
  Summary       - 전체 통합 요약 (% 지표 포함)
  네이버_vs_구글  - 플랫폼 비교
  Naver_SA      - 네이버 SA 상세
  Google_SA     - 구글 SA 상세
  Google_DA     - 구글 DA(PMax) 상세
  전환_상세      - 버튼별 전환 + 클릭 대비 %
  Naver_키워드   - 네이버 키워드 상세
  Google_키워드  - 구글 키워드 상세
  Google_검색어  - 구글 검색어 분석
  Raw           - 중앙 데이터 저장소
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from copy import copy
import pandas as pd
import numpy as np
from datetime import datetime


# ────────────────────────────────────────
# 스타일 상수
# ────────────────────────────────────────
C_NAVY   = '1E3A5F'
C_NAVER  = '1B6B3A'
C_GOOGLE = '4285F4'
C_DA     = 'C0392B'
C_WHITE  = 'FFFFFF'
C_LIGHT  = 'F2F6FB'
C_SUBHDR = '2B5797'
C_PCT    = 'FFF3CD'
C_PCT2   = 'E8F5E9'

NUM_INT  = '#,##0'
NUM_PCT  = '0.00%'
NUM_DEC  = '#,##0.0'
NUM_WON  = '#,##0"원"'

thin   = Side(style='thin', color='CCCCCC')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# 매체별 차트 색상
CHART_COLORS = {
    C_NAVER:  '1B6B3A',
    C_GOOGLE: '4285F4',
    C_DA:     'C0392B',
    C_NAVY:   '1E3A5F',
}

# ── 시트별 차트 너비 (칼럼폭 합계 기준, 1 char ≈ 0.187cm) ──
# 미디어 시트 A-I: [14,14,10,14,10,14,10,10,10] = 106 chars → 약 17cm (여백 포함 보수적 설정)
# 비교/요약 시트  A-K: 최대 120 chars → 약 19cm
# 전환 시트     A-G: [28,12,16,12,16,12,16] = 112 chars → 약 18cm
CHART_W_MEDIA   = 17.0   # 미디어 상세 시트 (Naver_SA, Google_SA, Google_DA)
CHART_W_COMPARE = 19.0   # 비교 / 요약 시트
CHART_W_CONV    = 18.0   # 전환 상세 시트
CHART_H         = 11.5   # 차트 높이 (공통)


# ────────────────────────────────────────
# 헬퍼
# ────────────────────────────────────────
# SNS 채널 색상
SNS_CHANNEL_COLORS = {
    '📱 인스타그램':  'E1306C',
    '🧵 쓰레드':     '333333',
    '📝 티스토리':   'FF6600',
    '💬 카카오톡채널': 'C8A800',
    '▶ 유튜브':      'CC0000',
}
C_SNS_HDR = '6C3483'   # 종합 장표 SNS 섹션 헤더 색 (보라)


# ────────────────────────────────────────
# SNS 관리대장 연동 헬퍼
# ────────────────────────────────────────

def _read_sns_summary(sns_path: str):
    """SNS 관리대장의 대시보드 요약 행을 읽어 반환.
    반환: [(채널, 지표, 최신값, 전일대비, 최대, 최소, 평균), ...] 또는 None
    채널명(이모지 포함 문자열)이 있는 행만 수집하여 정확히 5개 채널을 반환.
    """
    try:
        wb = openpyxl.load_workbook(sns_path, data_only=True)
        if '📊 대시보드' not in wb.sheetnames:
            return None
        ws = wb['📊 대시보드']
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            cell0 = row[0]
            if cell0 and isinstance(cell0, str) and cell0.strip() and \
               any(cell0.startswith(ch) for ch in ['📱', '🧵', '📝', '💬', '▶']):
                rows.append(list(row[:7]))
        return rows if rows else None
    except Exception:
        return None


def _copy_sns_sheet(sns_path: str, sheet_name: str, target_wb):
    """SNS 관리대장의 시트를 값(스냅샷) 전용으로 target_wb 에 복사."""
    try:
        src_wb = openpyxl.load_workbook(sns_path, data_only=True)
        if sheet_name not in src_wb.sheetnames:
            return
        src_ws = src_wb[sheet_name]
        tgt_ws = target_wb.create_sheet(sheet_name)
        tgt_ws.sheet_view.showGridLines = False

        # 셀 값 + 서식 복사
        for src_row in src_ws.iter_rows():
            for sc in src_row:
                tc = tgt_ws.cell(row=sc.row, column=sc.column, value=sc.value)
                if sc.has_style:
                    try:
                        tc.font        = copy(sc.font)
                        tc.fill        = copy(sc.fill)
                        tc.alignment   = copy(sc.alignment)
                        tc.border      = copy(sc.border)
                        tc.number_format = sc.number_format
                    except Exception:
                        pass

        # 병합 셀 복사
        for rng in src_ws.merged_cells.ranges:
            try:
                tgt_ws.merge_cells(str(rng))
            except Exception:
                pass

        # 열 너비 / 행 높이 복사
        for col_ltr, col_dim in src_ws.column_dimensions.items():
            tgt_ws.column_dimensions[col_ltr].width = col_dim.width
        for row_idx, row_dim in src_ws.row_dimensions.items():
            tgt_ws.row_dimensions[row_idx].height = row_dim.height

        # 스냅샷 안내 (A1 셀 위 삽입 대신 시트 탭 색으로 구분)
        channel = sheet_name
        tgt_ws.sheet_properties.tabColor = SNS_CHANNEL_COLORS.get(channel, '888888')

    except Exception:
        pass  # SNS 파일 오류는 무시하고 계속 진행


def _s(v): return 0.0 if (v is None or (isinstance(v, float) and np.isnan(v))) else float(v)

def _pct(num, den):
    try:
        d = float(den)
        return float(num) / d if d > 0 else 0.0
    except:
        return 0.0

def _hdr(ws, row, col, value, bg=C_NAVY, fg=C_WHITE, bold=True, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='맑은 고딕', bold=bold, color=fg, size=size)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER
    return c

def _cell(ws, row, col, value, fmt=None, bold=False, bg=None, align='center', color=None):
    c = ws.cell(row=row, column=col, value=value)
    fc = color if color else '000000'
    c.font = Font(name='맑은 고딕', bold=bold, size=10, color=fc)
    c.alignment = Alignment(horizontal=align, vertical='center')
    c.border = BORDER
    if fmt: c.number_format = fmt
    if bg:  c.fill = PatternFill('solid', fgColor=bg)
    return c

def _title(ws, row, col_span, text, bg=C_NAVY, size=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1, value=f'  {text}')
    c.font = Font(name='맑은 고딕', bold=True, size=size, color=C_WHITE)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(vertical='center')
    ws.row_dimensions[row].height = 26

def _section(ws, row, col_span, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1, value=f'  ■ {text}')
    c.font = Font(name='맑은 고딕', bold=True, size=11, color='1E3A5F')
    ws.row_dimensions[row].height = 20

def _write_table(ws, df, start_row, start_col, hdr_bg, col_fmts=None, col_bgs=None):
    col_fmts = col_fmts or {}
    col_bgs  = col_bgs  or {}
    for ci, col in enumerate(df.columns):
        _hdr(ws, start_row, start_col + ci, col, bg=hdr_bg)
    for ri, row in enumerate(df.itertuples(index=False), 1):
        stripe = C_LIGHT if ri % 2 == 0 else C_WHITE
        for ci, val in enumerate(row):
            col_name = df.columns[ci]
            bg  = col_bgs.get(col_name, stripe)
            fmt = col_fmts.get(col_name)
            _cell(ws, start_row + ri, start_col + ci, val, fmt=fmt, bg=bg)
    return start_row + len(df) + 1   # 마지막 데이터 행 다음 행

def _autowidth(ws, df, start_col=1):
    for i, col in enumerate(df.columns, start_col):
        try:
            data_max = int(df[col].astype(str).map(len).max())
        except:
            data_max = 10
        width = min(max(data_max, len(str(col))) + 2, 32)
        ws.column_dimensions[get_column_letter(i)].width = width

def _conv_dict(conv_df):
    if conv_df is None or conv_df.empty:
        return {}
    d = {}
    for _, r in conv_df.iterrows():
        key = str(r.get('전환유형', ''))
        val = float(str(r.get('전환수', 0)).replace(',', '') or 0)
        d[key] = d.get(key, 0) + val
    return d

def _section_title(ws, row, col_span, text, color):
    """차트 섹션 제목 (차트 바로 위)"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1, value=f'  ■ {text}')
    c.font = Font(name='맑은 고딕', bold=True, size=11, color=C_WHITE)
    c.fill = PatternFill('solid', fgColor=color)
    c.alignment = Alignment(vertical='center')
    ws.row_dimensions[row].height = 20


# ────────────────────────────────────────
# 차트 생성 함수
# ────────────────────────────────────────

def _data_labels(show_val=True, show_pct=False, show_cat=False, num_fmt='#,##0'):
    """데이터 레이블 객체 생성"""
    dl = DataLabelList()
    dl.showVal      = show_val
    dl.showPercent  = show_pct
    dl.showCatName  = show_cat
    dl.showSerName  = False
    dl.showLegendKey = False
    if num_fmt:
        dl.numFmt = num_fmt
    return dl


def _make_line_chart(ws, hdr_row, data_end_row, data_col, title, y_title,
                     color_hex, chart_width=CHART_W_MEDIA, fmt='#,##0'):
    """라인 차트 생성 (반환만, 배치는 호출자가)"""
    if data_end_row <= hdr_row:
        return None
    chart = LineChart()
    chart.title      = title
    chart.style      = 10
    chart.width      = chart_width
    chart.height     = CHART_H
    chart.smooth     = True
    chart.y_axis.numFmt = fmt
    chart.y_axis.title  = y_title
    chart.x_axis.title  = '날짜'

    data = Reference(ws, min_col=data_col, max_col=data_col,
                     min_row=hdr_row, max_row=data_end_row)
    chart.add_data(data, titles_from_data=True)

    try:
        s = chart.series[0]
        s.graphicalProperties.line.solidFill  = color_hex
        s.graphicalProperties.line.width      = 25000
        s.marker.symbol = 'circle'
        s.marker.size   = 5
        s.marker.graphicalProperties.solidFill            = color_hex
        s.marker.graphicalProperties.line.solidFill       = color_hex
    except Exception:
        pass

    # 수치 레이블 표시
    chart.dataLabels = _data_labels(show_val=True, num_fmt=fmt)

    cats = Reference(ws, min_col=1, max_col=1,
                     min_row=hdr_row + 1, max_row=data_end_row)
    chart.set_categories(cats)
    return chart


def _make_bar_chart(ws, hdr_row, data_end_row, data_cols, title,
                    colors=None, chart_width=CHART_W_COMPARE):
    """클러스터 바차트 생성"""
    if data_end_row <= hdr_row:
        return None
    chart = BarChart()
    chart.type     = 'col'
    chart.style    = 10
    chart.title    = title
    chart.width    = chart_width
    chart.height   = CHART_H
    chart.y_axis.numFmt = '#,##0'
    chart.gapWidth = 100

    default_colors = ['1B6B3A', '4285F4', 'C0392B', 'F4B400', '8E44AD']
    for i, col in enumerate(data_cols):
        data = Reference(ws, min_col=col, max_col=col,
                         min_row=hdr_row, max_row=data_end_row)
        chart.add_data(data, titles_from_data=True)
        try:
            c = colors[i] if colors and i < len(colors) else default_colors[i % len(default_colors)]
            chart.series[i].graphicalProperties.solidFill = c
        except Exception:
            pass

    # 수치 레이블 표시
    chart.dataLabels = _data_labels(show_val=True, num_fmt='#,##0')

    cats = Reference(ws, min_col=1, max_col=1,
                     min_row=hdr_row + 1, max_row=data_end_row)
    chart.set_categories(cats)
    return chart


def _make_pie_chart(ws, hdr_row, data_end_row, data_col, title,
                    chart_width=CHART_W_COMPARE):
    """파이 차트 생성"""
    if data_end_row <= hdr_row:
        return None
    chart = PieChart()
    chart.title  = title
    chart.style  = 10
    chart.width  = chart_width
    chart.height = CHART_H

    data = Reference(ws, min_col=data_col, max_col=data_col,
                     min_row=hdr_row, max_row=data_end_row)
    chart.add_data(data, titles_from_data=True)

    cats = Reference(ws, min_col=1, max_col=1,
                     min_row=hdr_row + 1, max_row=data_end_row)
    chart.set_categories(cats)

    # % + 카테고리명 + 수치 모두 표시
    chart.dataLabels = _data_labels(show_val=True, show_pct=True,
                                    show_cat=True, num_fmt='#,##0')
    return chart


def _place_chart(ws, chart, anchor_row, anchor_col=1):
    """차트를 지정 셀에 배치"""
    if chart is None:
        return
    anchor = f'{get_column_letter(anchor_col)}{anchor_row}'
    ws.add_chart(chart, anchor)


# ────────────────────────────────────────
# Summary 시트
# ────────────────────────────────────────
def build_summary(ws, raw_df, sa_conv, da_conv, period_label):
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False
    _title(ws, 1, 10, f'광고 성과 요약  |  {period_label}', size=14)
    ws.row_dimensions[2].height = 6

    row = 3
    sa_cd = _conv_dict(sa_conv)
    da_cd = _conv_dict(da_conv)

    grp = raw_df.groupby('매체').agg(
        노출=('노출','sum'), 클릭=('클릭','sum'),
        전환=('전환','sum'), 비용=('비용','sum')
    ).reset_index()

    total_노출 = grp['노출'].sum()
    total_클릭 = grp['클릭'].sum()
    total_비용 = grp['비용'].sum()

    sa_상세 = sa_cd.get('상세보기_버튼클릭', 0); sa_예약 = sa_cd.get('상담예약_버튼클릭', 0)
    sa_신청 = sa_cd.get('상담신청_버튼클릭', 0); da_상세 = da_cd.get('상세보기_버튼클릭', 0)
    da_예약 = da_cd.get('상담예약_버튼클릭', 0); da_신청 = da_cd.get('상담신청_버튼클릭', 0)

    # ── 매체 통합 요약 테이블 ──────────────────
    _section(ws, row, 10, '전체 매체 통합 요약')
    row += 1
    hdrs = ['매체','노출','클릭','CTR\n(클릭/노출)','CPC','비용',
            '노출비중','클릭비중','비용비중','전환(총)']
    PCT_COLS = {'CTR\n(클릭/노출)','노출비중','클릭비중','비용비중'}
    sum_hdr_row = row
    for ci, h in enumerate(hdrs):
        _hdr(ws, row, ci+1, h, bg=C_PCT if h in PCT_COLS else C_SUBHDR)
    row += 1

    for _, r in grp.iterrows():
        노출=_s(r['노출']); 클릭=_s(r['클릭']); 비용=_s(r['비용'])
        ctr = _pct(클릭, 노출)
        cpc = int(비용/클릭) if 클릭>0 else 0
        vals = [r['매체'], int(노출), int(클릭), ctr, cpc, int(비용),
                _pct(노출,total_노출), _pct(클릭,total_클릭), _pct(비용,total_비용),
                round(_s(r['전환']),1)]
        fmts = [None,NUM_INT,NUM_INT,NUM_PCT,NUM_INT,NUM_INT,NUM_PCT,NUM_PCT,NUM_PCT,NUM_DEC]
        bgs  = [None,None,None,C_PCT,None,None,C_PCT,C_PCT,C_PCT,None]
        stripe = C_LIGHT if row % 2 == 0 else C_WHITE
        for ci, (v, f, b) in enumerate(zip(vals, fmts, bgs)):
            _cell(ws, row, ci+1, v, fmt=f, bg=b or stripe)
        row += 1
    sum_data_end = row - 1
    row += 1

    # ── 버튼 전환율 블록 ──────────────────────
    _section(ws, row, 10, '클릭 대비 버튼 전환율 (구글)')
    row += 1
    conv_hdrs = ['구분','상세보기\n클릭수','상세보기율\n(클릭대비)',
                 '상담예약\n클릭수','상담예약율\n(클릭대비)',
                 '상담신청\n클릭수','상담신청율\n(클릭대비)',
                 '버튼클릭\n합계','버튼클릭\n총전환율']
    for ci, h in enumerate(conv_hdrs):
        _hdr(ws, row, ci+1, h, bg=C_PCT if '율' in h else C_SUBHDR)
    row += 1

    sa_클릭 = raw_df[raw_df['매체']=='Google_SA']['클릭'].sum()
    da_클릭 = raw_df[raw_df['매체']=='Google_DA']['클릭'].sum()
    total_g클릭 = sa_클릭 + da_클릭

    rows_conv = [
        ('Google SA', sa_클릭, sa_상세, sa_예약, sa_신청),
        ('Google DA', da_클릭, da_상세, da_예약, da_신청),
        ('Google 합계', total_g클릭, sa_상세+da_상세, sa_예약+da_예약, sa_신청+da_신청),
    ]
    for ri, (lbl, clk, 상세, 예약, 신청) in enumerate(rows_conv):
        합계 = 상세 + 예약 + 신청
        stripe = C_LIGHT if ri % 2 == 0 else C_WHITE
        vals = [lbl, int(상세), _pct(상세,clk), int(예약), _pct(예약,clk),
                int(신청), _pct(신청,clk), int(합계), _pct(합계,clk)]
        fmts = [None,NUM_INT,NUM_PCT,NUM_INT,NUM_PCT,NUM_INT,NUM_PCT,NUM_INT,NUM_PCT]
        bgs  = [None,None,C_PCT,None,C_PCT,None,C_PCT,None,C_PCT2]
        for ci, (v, f, b) in enumerate(zip(vals, fmts, bgs)):
            _cell(ws, row, ci+1, v, fmt=f, bg=b or stripe, bold=(ri==2))
        row += 1

    # 스타일
    ws.column_dimensions['A'].width = 16
    for col in ['B','C','D','E','F','G','H','I','J']:
        ws.column_dimensions[col].width = 12
    ws.freeze_panes = 'A3'
    ws.row_dimensions[1].height = 28

    # ── 차트 섹션 (테이블 아래 A열 기준 정렬) ────
    row += 1
    _section_title(ws, row, 10, '매체별 비용 비중', C_NAVY)
    row += 1
    chart_pie = _make_pie_chart(ws, sum_hdr_row, sum_data_end, data_col=6,
                                title='매체별 비용 비중',
                                chart_width=CHART_W_COMPARE)
    _place_chart(ws, chart_pie, anchor_row=row, anchor_col=1)

    # 파이 차트 높이(약 22행) 아래 바 차트
    row += 22
    _section_title(ws, row, 10, '매체별 클릭수 비교', C_NAVY)
    row += 1
    chart_bar = _make_bar_chart(ws, sum_hdr_row, sum_data_end,
                                data_cols=[3],
                                title='매체별 클릭수 비교',
                                colors=['4285F4'],
                                chart_width=CHART_W_COMPARE)
    _place_chart(ws, chart_bar, anchor_row=row, anchor_col=1)


# ────────────────────────────────────────
# 비교 시트
# ────────────────────────────────────────
def build_comparison(ws, raw_df, sa_conv, da_conv):
    ws.title = '네이버_vs_구글'
    ws.sheet_view.showGridLines = False
    _title(ws, 1, 11, '네이버 vs 구글 비교')

    sa_cd = _conv_dict(sa_conv)
    da_cd = _conv_dict(da_conv)
    naver  = raw_df[raw_df['매체'].str.startswith('Naver')]
    google = raw_df[raw_df['매체'].str.startswith('Google')]

    total_노출 = raw_df['노출'].sum()
    total_클릭 = raw_df['클릭'].sum()
    total_비용 = raw_df['비용'].sum()

    hdrs = ['구분','노출','클릭','CTR','CPC','비용',
            '노출비중','클릭비중','비용비중','버튼클릭\n합계','버튼클릭\n전환율']
    PCT = {'CTR','노출비중','클릭비중','비용비중','버튼클릭\n전환율'}
    for ci, h in enumerate(hdrs):
        _hdr(ws, 2, ci+1, h, bg=C_PCT if h in PCT else C_SUBHDR)

    g_btn = (sa_cd.get('상세보기_버튼클릭',0) + sa_cd.get('상담예약_버튼클릭',0) +
             sa_cd.get('상담신청_버튼클릭',0) + da_cd.get('상세보기_버튼클릭',0) +
             da_cd.get('상담예약_버튼클릭',0) + da_cd.get('상담신청_버튼클릭',0))

    # 네이버: 별도 전환파일 없음 → raw 데이터의 '총 전환수' 사용
    naver_btn = int(naver['전환'].sum())

    comp_hdr_row = 2
    for ri, (lbl, grp, color, btn) in enumerate([
        ('네이버 SA', naver, C_NAVER, naver_btn),
        ('구글 광고', google, C_GOOGLE, g_btn),
    ]):
        노출=grp['노출'].sum(); 클릭=grp['클릭'].sum(); 비용=grp['비용'].sum()
        ctr = _pct(클릭, 노출)
        cpc = int(비용/클릭) if 클릭>0 else 0
        btn_rate = _pct(btn, 클릭)
        vals = [lbl, int(노출), int(클릭), ctr, cpc, int(비용),
                _pct(노출,total_노출), _pct(클릭,total_클릭), _pct(비용,total_비용),
                int(btn), btn_rate]
        fmts = [None,NUM_INT,NUM_INT,NUM_PCT,NUM_INT,NUM_INT,
                NUM_PCT,NUM_PCT,NUM_PCT,NUM_INT,NUM_PCT]
        row = 3 + ri
        for ci, (v, f) in enumerate(zip(vals, fmts)):
            c = _cell(ws, row, ci+1, v, fmt=f)
            if ci == 0:
                c.font = Font(name='맑은 고딕', bold=True, color=C_WHITE, size=10)
                c.fill = PatternFill('solid', fgColor=color)
            elif f == NUM_PCT:
                c.fill = PatternFill('solid', fgColor=C_PCT)
    comp_data_end = 4

    ws.row_dimensions[2].height = 30
    for col, w in zip('ABCDEFGHIJK', [14,14,10,8,10,14,10,10,10,10,10]):
        ws.column_dimensions[col].width = w

    # ── 차트: 노출 / 클릭 / 비용 비교 (테이블 바로 아래) ──
    chart_row = comp_data_end + 2
    _section_title(ws, chart_row, 11, '노출 / 클릭 / 비용 매체 비교', C_NAVY)
    chart_row += 1
    chart = _make_bar_chart(ws, comp_hdr_row, comp_data_end,
                            data_cols=[2, 3, 6],
                            title='노출 / 클릭 / 비용 매체 비교',
                            colors=[C_NAVER, C_GOOGLE, 'F4B400'],
                            chart_width=CHART_W_COMPARE)
    _place_chart(ws, chart, anchor_row=chart_row, anchor_col=1)


# ────────────────────────────────────────
# 매체별 상세 시트
# ────────────────────────────────────────
def build_media_sheet(ws, raw_df, media_prefix, title, color, conv_df=None):
    ws.title = title
    ws.sheet_view.showGridLines = False
    _title(ws, 1, 9, f'{title} 상세 리포트', bg=color)

    df = raw_df[raw_df['매체'].str.startswith(media_prefix)].copy()
    conv_cd   = _conv_dict(conv_df)
    total_클릭 = df['클릭'].sum()

    # ── 버튼 전환 요약 ───────────────────────
    row = 3
    _section(ws, row, 9, '기간 버튼 전환 요약')
    row += 1
    btn_hdrs = ['상세보기\n클릭수','상세보기율','상담예약\n클릭수','상담예약율',
                '상담신청\n클릭수','상담신청율','버튼클릭\n합계','버튼클릭\n총전환율','(기준: 전체클릭)']
    for ci, h in enumerate(btn_hdrs):
        _hdr(ws, row, ci+1, h, bg=C_PCT if '율' in h else color)
    row += 1
    상세=conv_cd.get('상세보기_버튼클릭',0); 예약=conv_cd.get('상담예약_버튼클릭',0)
    신청=conv_cd.get('상담신청_버튼클릭',0)
    # 네이버: 별도 전환파일 없음 → raw 데이터 '총 전환수' 사용
    if conv_df is None:
        합계 = int(df['전환'].sum())
    else:
        합계 = 상세 + 예약 + 신청
    vals_b = [int(상세),_pct(상세,total_클릭),int(예약),_pct(예약,total_클릭),
              int(신청),_pct(신청,total_클릭),int(합계),_pct(합계,total_클릭),
              f'{int(total_클릭):,}클릭 기준']
    fmts_b = [NUM_INT,NUM_PCT,NUM_INT,NUM_PCT,NUM_INT,NUM_PCT,NUM_INT,NUM_PCT,None]
    bgs_b  = [None,C_PCT,None,C_PCT,None,C_PCT,None,C_PCT2,C_LIGHT]
    for ci, (v, f, b) in enumerate(zip(vals_b, fmts_b, bgs_b)):
        _cell(ws, row, ci+1, v, fmt=f, bg=b or C_WHITE, bold=(ci==7))
    row += 2

    # ── 일별 성과 ────────────────────────────
    _section(ws, row, 9, '일별 성과')
    row += 1
    daily = df.groupby('날짜', as_index=False).agg(
        노출=('노출','sum'), 클릭=('클릭','sum'), 비용=('비용','sum')
    ).sort_values('날짜')
    daily['날짜']            = pd.to_datetime(daily['날짜']).dt.strftime('%Y-%m-%d')
    daily['CTR(클릭/노출)'] = daily.apply(lambda r: _pct(r['클릭'],r['노출']), axis=1)
    daily['CPC']             = daily.apply(lambda r: int(r['비용']/r['클릭']) if r['클릭']>0 else 0, axis=1)
    daily['클릭비중(일별)'] = daily.apply(lambda r: _pct(r['클릭'],total_클릭), axis=1)

    disp = daily[['날짜','노출','클릭','CTR(클릭/노출)','CPC','비용','클릭비중(일별)']]
    fmts_d = {'노출':NUM_INT,'클릭':NUM_INT,'CTR(클릭/노출)':NUM_PCT,
              'CPC':NUM_INT,'비용':NUM_INT,'클릭비중(일별)':NUM_PCT}
    bgs_d  = {'CTR(클릭/노출)':C_PCT,'클릭비중(일별)':C_PCT}

    daily_hdr_row = row                     # 헤더 행 (차트 데이터 참조용)
    row = _write_table(ws, disp, row, 1, color, fmts_d, bgs_d)
    daily_data_end = daily_hdr_row + len(disp)
    row += 1

    # ── 디바이스별 요약 ──────────────────────
    _section(ws, row, 9, '디바이스별 요약')
    row += 1
    dev = df.groupby('디바이스', as_index=False).agg(
        노출=('노출','sum'), 클릭=('클릭','sum'), 비용=('비용','sum')
    )
    dev['CTR(클릭/노출)'] = dev.apply(lambda r: _pct(r['클릭'],r['노출']), axis=1)
    dev['CPC']            = dev.apply(lambda r: int(r['비용']/r['클릭']) if r['클릭']>0 else 0, axis=1)
    dev['클릭비중']       = dev.apply(lambda r: _pct(r['클릭'],total_클릭), axis=1)
    dev['노출비중']       = dev.apply(lambda r: _pct(r['노출'],df['노출'].sum()), axis=1)
    dev['비용비중']       = dev.apply(lambda r: _pct(r['비용'],df['비용'].sum()), axis=1)

    disp2 = dev[['디바이스','노출','클릭','CTR(클릭/노출)','CPC','비용','노출비중','클릭비중','비용비중']]
    fmts2 = {'노출':NUM_INT,'클릭':NUM_INT,'CTR(클릭/노출)':NUM_PCT,
             'CPC':NUM_INT,'비용':NUM_INT,'노출비중':NUM_PCT,'클릭비중':NUM_PCT,'비용비중':NUM_PCT}
    bgs2  = {'CTR(클릭/노출)':C_PCT,'노출비중':C_PCT,'클릭비중':C_PCT,'비용비중':C_PCT}
    row = _write_table(ws, disp2, row, 1, color, fmts2, bgs2)

    # 칼럼 너비 확정 (A~I)
    for col, w in zip('ABCDEFGHI', [14,14,10,14,10,14,10,10,10]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A4'

    # ── 차트 섹션 (디바이스 테이블 바로 아래, A열 기준) ──
    chart_row = row + 1

    _section_title(ws, chart_row, 9, '일별 클릭수 추이', color)
    chart_row += 1
    c1 = _make_line_chart(ws, daily_hdr_row, daily_data_end,
                          data_col=3, title='일별 클릭수 추이',
                          y_title='클릭수', color_hex=color,
                          chart_width=CHART_W_MEDIA)
    _place_chart(ws, c1, anchor_row=chart_row, anchor_col=1)

    # 클릭 차트 높이(≈22행) 아래 비용 차트
    chart_row += 22
    _section_title(ws, chart_row, 9, '일별 비용 추이', color)
    chart_row += 1
    c2 = _make_line_chart(ws, daily_hdr_row, daily_data_end,
                          data_col=6, title='일별 비용 추이',
                          y_title='비용(원)', color_hex=color,
                          chart_width=CHART_W_MEDIA)
    _place_chart(ws, c2, anchor_row=chart_row, anchor_col=1)


# ────────────────────────────────────────
# 전환_상세 시트
# ────────────────────────────────────────
def build_conversion_sheet(ws, sa_conv, da_conv, raw_df):
    ws.title = '전환_상세'
    ws.sheet_view.showGridLines = False
    _title(ws, 1, 7, '구글 버튼 전환 상세 (SA / DA 분리)', bg=C_GOOGLE)

    sa_클릭    = raw_df[raw_df['매체']=='Google_SA']['클릭'].sum()
    da_클릭    = raw_df[raw_df['매체']=='Google_DA']['클릭'].sum()
    total_클릭 = sa_클릭 + da_클릭

    row = 3
    hdrs = ['전환 유형','SA 전환수','SA 전환율\n(SA클릭 대비)',
            'DA 전환수','DA 전환율\n(DA클릭 대비)',
            '합계 전환수','합계 전환율\n(전체클릭 대비)']
    for ci, h in enumerate(hdrs):
        _hdr(ws, row, ci+1, h, bg=C_PCT if '율' in h else C_SUBHDR)
    conv_hdr_row = row
    row += 1

    sa_cd = _conv_dict(sa_conv)
    da_cd = _conv_dict(da_conv)
    유형_list  = ['페이지조회','상세보기_버튼클릭','상담예약_버튼클릭','상담신청_버튼클릭']
    유형_label = ['페이지 조회','상세보기 버튼클릭','상담예약 버튼클릭','상담신청 버튼클릭']

    for ri, (key, label) in enumerate(zip(유형_list, 유형_label)):
        sa_v = sa_cd.get(key, 0); da_v = da_cd.get(key, 0); total_v = sa_v + da_v
        stripe = C_LIGHT if ri % 2 == 0 else C_WHITE
        vals = [label, int(sa_v), _pct(sa_v,sa_클릭),
                int(da_v), _pct(da_v,da_클릭),
                int(total_v), _pct(total_v,total_클릭)]
        fmts = [None,NUM_INT,NUM_PCT,NUM_INT,NUM_PCT,NUM_INT,NUM_PCT]
        bgs  = [None,None,C_PCT,None,C_PCT,None,C_PCT2]
        for ci, (v, f, b) in enumerate(zip(vals, fmts, bgs)):
            _cell(ws, row, ci+1, v, fmt=f, bg=b or stripe, bold=(key!='페이지조회'))
        row += 1
    conv_data_end = row - 1

    # 합계 행
    row += 1
    btn_keys = ['상세보기_버튼클릭','상담예약_버튼클릭','상담신청_버튼클릭']
    sa_btn = sum(sa_cd.get(k,0) for k in btn_keys)
    da_btn = sum(da_cd.get(k,0) for k in btn_keys)
    total_btn = sa_btn + da_btn
    for ci, (v, f) in enumerate(zip(
        ['버튼클릭 합계 (페이지조회 제외)',
         int(sa_btn), _pct(sa_btn,sa_클릭), int(da_btn), _pct(da_btn,da_클릭),
         int(total_btn), _pct(total_btn,total_클릭)],
        [None,NUM_INT,NUM_PCT,NUM_INT,NUM_PCT,NUM_INT,NUM_PCT]
    )):
        c = _cell(ws, row, ci+1, v, fmt=f, bold=True, bg=C_NAVY if ci==0 else C_PCT2)
        if ci == 0:
            c.font = Font(name='맑은 고딕', bold=True, color=C_WHITE, size=10)

    row += 2
    ws.cell(row=row, column=1,
            value=f'※ SA클릭: {int(sa_클릭):,}  |  DA클릭: {int(da_클릭):,}  |  전체구글클릭: {int(total_클릭):,}'
            ).font = Font(name='맑은 고딕', size=9, color='666666', italic=True)

    ws.column_dimensions['A'].width = 28
    for col, w in zip('BCDEFG', [12,16,12,16,12,16]):
        ws.column_dimensions[col].width = w

    # ── 차트: SA vs DA 버튼 전환 비교 (합계 행 아래) ──
    chart_row = row + 2
    _section_title(ws, chart_row, 7, 'SA vs DA 버튼 전환 비교', C_GOOGLE)
    chart_row += 1
    chart = _make_bar_chart(ws, conv_hdr_row, conv_data_end,
                            data_cols=[2, 4],
                            title='SA vs DA 버튼 전환 비교',
                            colors=[C_GOOGLE, C_DA],
                            chart_width=CHART_W_CONV)
    _place_chart(ws, chart, anchor_row=chart_row, anchor_col=1)


# ────────────────────────────────────────
# 키워드 시트
# ────────────────────────────────────────
def build_keyword_sheet(ws, kw_df, title, color):
    ws.title = title[:31]
    ws.sheet_view.showGridLines = False
    _title(ws, 1, 9, title, bg=color)

    df = kw_df.copy()
    total_클릭 = df['클릭'].sum() if '클릭' in df.columns else 1
    비용_col = '광고비(VAT별도)' if '광고비(VAT별도)' in df.columns else ('비용' if '비용' in df.columns else None)

    if '노출' in df.columns and '클릭' in df.columns:
        df['CTR(클릭/노출)'] = df.apply(lambda r: _pct(r['클릭'],r['노출']), axis=1)
        df['클릭비중']       = df.apply(lambda r: _pct(r['클릭'],total_클릭), axis=1)
    if 비용_col:
        total_비용 = df[비용_col].sum()
        df['비용비중'] = df.apply(lambda r: _pct(r[비용_col],total_비용), axis=1)

    fmts = {'노출':NUM_INT,'클릭':NUM_INT,'CTR(클릭/노출)':NUM_PCT,
            '클릭비중':NUM_PCT,'비용비중':NUM_PCT,'광고비(VAT별도)':NUM_INT,
            '비용':NUM_INT,'전환':NUM_DEC,'평균노출순위':NUM_DEC,'평균CPC':NUM_INT}
    bgs  = {'CTR(클릭/노출)':C_PCT,'클릭비중':C_PCT,'비용비중':C_PCT}

    _write_table(ws, df, 2, 1, color, fmts, bgs)
    _autowidth(ws, df)
    ws.freeze_panes = 'A3'


# ────────────────────────────────────────
# 검색어 시트
# ────────────────────────────────────────
def build_search_terms(ws, st_df):
    ws.title = 'Google_검색어'
    ws.sheet_view.showGridLines = False
    _title(ws, 1, 8, '구글 검색어 분석', bg=C_GOOGLE)

    df = st_df.copy()
    total_클릭 = df['클릭'].sum() if '클릭' in df.columns else 1
    if '클릭' in df.columns and '노출' in df.columns:
        df['CTR(클릭/노출)'] = df.apply(lambda r: _pct(r['클릭'],r['노출']), axis=1)
        df['클릭비중']       = df.apply(lambda r: _pct(r['클릭'],total_클릭), axis=1)

    fmts = {'노출':NUM_INT,'클릭':NUM_INT,'CTR(클릭/노출)':NUM_PCT,
            '클릭비중':NUM_PCT,'평균CPC':NUM_INT,'비용':NUM_INT,'전환':NUM_DEC}
    bgs  = {'CTR(클릭/노출)':C_PCT,'클릭비중':C_PCT}

    _write_table(ws, df, 2, 1, C_GOOGLE, fmts, bgs)
    _autowidth(ws, df)
    ws.freeze_panes = 'A3'


# ────────────────────────────────────────
# 종합 장표 시트 (첫 번째 시트)
# ────────────────────────────────────────
def build_overview_sheet(ws, raw_df, sa_conv, da_conv, period_label, sns_rows=None):
    ws.title = '📋 종합 장표'
    ws.sheet_view.showGridLines = False

    _title(ws, 1, 12, f'종합 마케팅 성과 장표  |  {period_label}', bg=C_NAVY, size=14)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 8

    sa_cd = _conv_dict(sa_conv)
    da_cd = _conv_dict(da_conv)

    naver_df = raw_df[raw_df['매체'].str.startswith('Naver')]
    gsa_df   = raw_df[raw_df['매체'] == 'Google_SA']
    gda_df   = raw_df[raw_df['매체'] == 'Google_DA']

    BTN_KEYS = ['상세보기_버튼클릭', '상담예약_버튼클릭', '상담신청_버튼클릭']

    def _ad_row(label, df, conv_cd_or_none):
        노출 = df['노출'].sum(); 클릭 = df['클릭'].sum(); 비용 = df['비용'].sum()
        ctr  = _pct(클릭, 노출)
        cpc  = int(비용 / 클릭) if 클릭 > 0 else 0
        if conv_cd_or_none is None:
            btn = int(df['전환'].sum())        # 네이버: 총 전환수
        else:
            btn = int(sum(conv_cd_or_none.get(k, 0) for k in BTN_KEYS))
        return [label, int(노출), int(클릭), ctr, cpc, int(비용), btn, _pct(btn, 클릭)]

    # ── 광고 성과 섹션 ────────────────────
    row = 3
    _section(ws, row, 12, '광고 성과 요약 (기간 합계)')
    row += 1

    AD_HDRS = ['매체', '노출', '클릭', 'CTR', 'CPC', '광고비', '버튼 전환', '버튼전환율']
    PCT_AD  = {'CTR', '버튼전환율'}
    for ci, h in enumerate(AD_HDRS):
        _hdr(ws, row, ci+1, h, bg=C_PCT if h in PCT_AD else C_SUBHDR)
    row += 1

    ad_data = [
        ('🟢 네이버 SA', naver_df, None,  C_NAVER),
        ('🔵 구글 SA',   gsa_df,   sa_cd, C_GOOGLE),
        ('🔴 구글 DA',   gda_df,   da_cd, C_DA),
    ]
    AD_FMTS = [None, NUM_INT, NUM_INT, NUM_PCT, NUM_INT, NUM_INT, NUM_INT, NUM_PCT]

    for ri, (lbl, df, cd, color) in enumerate(ad_data):
        rdata  = _ad_row(lbl, df, cd)
        stripe = C_LIGHT if ri % 2 == 0 else C_WHITE
        for ci, (v, f) in enumerate(zip(rdata, AD_FMTS)):
            c = _cell(ws, row, ci+1, v, fmt=f, bg=stripe if ci > 0 else None)
            if ci == 0:
                c.font = Font(name='맑은 고딕', bold=True, color=C_WHITE, size=10)
                c.fill = PatternFill('solid', fgColor=color)
            elif f == NUM_PCT:
                c.fill = PatternFill('solid', fgColor=C_PCT)
        row += 1

    # 전체 합계 행
    total_노출 = raw_df['노출'].sum(); total_클릭 = raw_df['클릭'].sum()
    total_비용 = raw_df['비용'].sum()
    naver_btn  = int(naver_df['전환'].sum())
    sa_btn     = int(sum(sa_cd.get(k, 0) for k in BTN_KEYS))
    da_btn     = int(sum(da_cd.get(k, 0) for k in BTN_KEYS))
    total_btn  = naver_btn + sa_btn + da_btn

    total_row = ['📊 전체 합계', int(total_노출), int(total_클릭),
                 _pct(total_클릭, total_노출),
                 int(total_비용 / total_클릭) if total_클릭 > 0 else 0,
                 int(total_비용), total_btn, _pct(total_btn, total_클릭)]
    for ci, (v, f) in enumerate(zip(total_row, AD_FMTS)):
        c = _cell(ws, row, ci+1, v, fmt=f, bold=True, bg=C_LIGHT)
        if ci == 0:
            c.font = Font(name='맑은 고딕', bold=True, color=C_WHITE, size=10)
            c.fill = PatternFill('solid', fgColor=C_NAVY)
        elif f == NUM_PCT:
            c.fill = PatternFill('solid', fgColor=C_PCT)
    row += 2

    # ── SNS 채널 현황 섹션 ────────────────
    _section(ws, row, 12, 'SNS 채널 현황 (최신 입력 데이터 기준)')
    row += 1

    SNS_HDRS = ['채널', '주요 지표', '최신값', '전일 대비 증감', '기간 최대', '기간 최소', '기간 평균']
    for ci, h in enumerate(SNS_HDRS):
        _hdr(ws, row, ci+1, h, bg=C_SNS_HDR)
    row += 1

    if sns_rows:
        for ri, rdata in enumerate(sns_rows):
            stripe  = C_LIGHT if ri % 2 == 0 else C_WHITE
            channel = str(rdata[0]) if rdata[0] else ''
            color   = SNS_CHANNEL_COLORS.get(channel, C_NAVY)
            for ci, v in enumerate(rdata[:7]):
                disp = v if v is not None else '-'
                c = _cell(ws, row, ci+1, disp, bg=stripe if ci > 0 else None)
                if ci == 0:
                    c.font = Font(name='맑은 고딕', bold=True, color=C_WHITE, size=10)
                    c.fill = PatternFill('solid', fgColor=color)
            row += 1
    else:
        # SNS 파일 미연결 안내
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1,
                    value='  ※ SNS 관리대장 파일을 함께 제공하면 최신 채널 현황이 자동으로 표시됩니다.')
        c.font = Font(name='맑은 고딕', size=10, color='888888', italic=True)
        c.fill = PatternFill('solid', fgColor='FAFAFA')
        row += 1

    # 참고 안내
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    note = ws.cell(row=row, column=1,
                   value='  ※ SNS 채널 상세 트렌드 및 차트는 하단 SNS 시트 탭 또는 SNS_채널_관리대장.xlsx 를 확인하세요.')
    note.font = Font(name='맑은 고딕', size=9, color='666666', italic=True)

    # 열 너비
    col_widths = [22, 14, 12, 10, 12, 14, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'


# ────────────────────────────────────────
# Raw 시트
# ────────────────────────────────────────
def build_raw(ws, raw_df):
    ws.title = 'Raw'
    df = raw_df.copy()
    df['날짜'] = pd.to_datetime(df['날짜']).dt.strftime('%Y-%m-%d')
    df['월']   = pd.to_datetime(df['월']).dt.strftime('%Y-%m')
    fmts = {'노출':NUM_INT,'클릭':NUM_INT,'전환':NUM_DEC,'비용':NUM_INT}
    _write_table(ws, df, 1, 1, C_NAVY, fmts)
    for col, w in zip('ABCDEFGHI', [10,10,12,16,8,14,10,10,14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'


# ────────────────────────────────────────
# 메인 빌더
# ────────────────────────────────────────
def build_report(raw_df, naver_kw_df=None, google_sa_kw_df=None,
                 google_da_kw_df=None, sa_conv_df=None, da_conv_df=None,
                 search_terms_df=None, period_label='', output_path='report.xlsx',
                 sns_tracker_path=None):
    """광고 성과 Excel 장표 생성.

    sns_tracker_path 를 지정하면:
      - 첫 번째 시트로 '📋 종합 장표' 가 추가됩니다 (광고 KPI + SNS 최신 현황)
      - SNS 채널 시트 5개가 스냅샷으로 마지막에 추가됩니다
    """

    # ── SNS 데이터 읽기 (선택) ──────────────
    sns_rows = None
    if sns_tracker_path:
        sns_rows = _read_sns_summary(sns_tracker_path)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 시트 생성 ──────────────────────────
    # 종합 장표를 항상 첫 번째로 (SNS 파일 없어도 광고 요약은 표시)
    ws_ov   = wb.create_sheet('📋 종합 장표')
    ws_sum  = wb.create_sheet('Summary')
    ws_cmp  = wb.create_sheet('네이버_vs_구글')
    ws_nav  = wb.create_sheet('Naver_SA')
    ws_gsa  = wb.create_sheet('Google_SA')
    ws_gda  = wb.create_sheet('Google_DA')
    ws_conv = wb.create_sheet('전환_상세')
    ws_nkw  = wb.create_sheet('Naver_키워드')
    ws_gkw  = wb.create_sheet('Google_키워드')
    ws_gst  = wb.create_sheet('Google_검색어')
    ws_raw  = wb.create_sheet('Raw')

    # ── 광고 시트 빌드 ─────────────────────
    build_overview_sheet(ws_ov, raw_df, sa_conv_df, da_conv_df, period_label, sns_rows)
    build_summary(ws_sum,  raw_df, sa_conv_df, da_conv_df, period_label)
    build_comparison(ws_cmp, raw_df, sa_conv_df, da_conv_df)
    build_media_sheet(ws_nav, raw_df, 'Naver',     'Naver_SA',  C_NAVER,  None)
    build_media_sheet(ws_gsa, raw_df, 'Google_SA', 'Google_SA', C_GOOGLE, sa_conv_df)
    build_media_sheet(ws_gda, raw_df, 'Google_DA', 'Google_DA', C_DA,     da_conv_df)

    if (sa_conv_df is not None and not sa_conv_df.empty) or \
       (da_conv_df is not None and not da_conv_df.empty):
        build_conversion_sheet(ws_conv, sa_conv_df, da_conv_df, raw_df)
    else:
        ws_conv['A1'] = '전환 데이터 없음'

    if naver_kw_df is not None and not naver_kw_df.empty:
        build_keyword_sheet(ws_nkw, naver_kw_df, 'Naver_키워드 상세', C_NAVER)
    else:
        ws_nkw['A1'] = '네이버 키워드 데이터 없음'

    g_kws = [d for d in [google_sa_kw_df, google_da_kw_df] if d is not None and not d.empty]
    if g_kws:
        build_keyword_sheet(ws_gkw, pd.concat(g_kws, ignore_index=True), 'Google_키워드 상세', C_GOOGLE)
    else:
        ws_gkw['A1'] = '구글 키워드 데이터 없음'

    if search_terms_df is not None and not search_terms_df.empty:
        build_search_terms(ws_gst, search_terms_df)
    else:
        ws_gst['A1'] = '검색어 데이터 없음'

    build_raw(ws_raw, raw_df)

    # ── SNS 시트 스냅샷 추가 (선택) ─────────
    if sns_tracker_path:
        SNS_SHEETS = ['📊 대시보드', '📱 인스타그램', '🧵 쓰레드',
                      '📝 티스토리', '💬 카카오톡채널', '▶ 유튜브']
        for sname in SNS_SHEETS:
            _copy_sns_sheet(sns_tracker_path, sname, wb)

    wb.save(output_path)
    print(f'✅ 장표 저장: {output_path}')
    return output_path
